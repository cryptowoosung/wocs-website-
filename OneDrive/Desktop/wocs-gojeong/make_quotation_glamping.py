#!/usr/bin/env python3
"""
WOCS 글램핑 텐트 전용 견적서 생성기
────────────────────────────────────────────────────────────
사용법:
  python3 make_quotation.py                    # 대화형 입력
  python3 make_quotation.py --type awning      # 천막·어닝 견적서
  python3 make_quotation.py --type glamping    # 글램핑 텐트 견적서
  python3 make_quotation.py --list             # 저장된 견적 목록 출력
────────────────────────────────────────────────────────────
"""

import os, sys, json, argparse
from datetime import date, timedelta
from pathlib import Path

# ── 의존성 체크 ──
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl --break-system-packages")
    sys.exit(1)

# PDF 변환 (선택 — 없어도 xlsx 생성 가능)
try:
    import subprocess
    HAS_LIBREOFFICE = subprocess.run(
        ['which', 'libreoffice'], capture_output=True).returncode == 0
except:
    HAS_LIBREOFFICE = False

# ── 경로 설정 ──
BASE_DIR    = Path(__file__).parent
OUTPUT_DIR  = BASE_DIR / 'output'
COUNTER_FILE = BASE_DIR / '.quote_counter.json'
OUTPUT_DIR.mkdir(exist_ok=True)

# ══════════════════════════════════════════════════════════
# 색상 팔레트 (고급 포레스트 그린 + 샴페인 골드)
# ══════════════════════════════════════════════════════════
C = {
    'DG'  : '1A3526',  # 딥 포레스트 그린 (헤더)
    'GOLD': 'B59035',  # 샴페인 골드 (합계)
    'SAGE': 'E8F0EA',  # 세이지 그린 (섹션)
    'CRM' : 'FDF8EE',  # 아이보리 크림 (입력)
    'LGRY': 'F4F6F4',  # 그린틴트 (짝수행)
    'W'   : 'FFFFFF',
    'DK'  : '1A1A1A',
    'GTXT': 'F0D080',  # 소프트 골드 텍스트
    'MDGN': '2D5A3D',  # 섹션 제목 텍스트
    'LGOL': 'FDF3DC',  # 합계값 배경
}

# ══════════════════════════════════════════════════════════
# 스타일 헬퍼
# ══════════════════════════════════════════════════════════
def fl(c):
    return PatternFill('solid', start_color=c, end_color=c)

def fn(sz=10, bold=False, color=None):
    return Font(name='맑은 고딕', size=sz, bold=bold, color=color or C['DK'])

def bd():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def ca(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def la(i=1): return Alignment(horizontal='left',  vertical='center', wrap_text=True, indent=i)
def ra(): return Alignment(horizontal='right', vertical='center', wrap_text=False)

def sc(ws, addr, val=None, font=None, fill=None, align=None, border=None, fmt=None):
    """set_cell — 셀 속성 한 번에 설정"""
    c = ws[addr]
    if val   is not None: c.value  = val
    if font  is not None: c.font   = font
    if fill  is not None: c.fill   = fill
    if align is not None: c.alignment = align
    if border is not None: c.border = border
    if fmt   is not None: c.number_format = fmt

def mc(ws, rng):
    """병합 — 이미 병합된 경우 해제 후 재병합"""
    try: ws.unmerge_cells(rng)
    except: pass
    ws.merge_cells(rng)

def apply_merge_border(ws, rng):
    """병합 범위 전체에 방향별 테두리 적용"""
    from openpyxl.utils import range_boundaries
    try: c1, r1, c2, r2 = range_boundaries(rng)
    except: return
    s = Side(style='thin'); n = Side(style=None)
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = Border(
                left   = s if c==c1 else n,
                right  = s if c==c2 else n,
                top    = s if r==r1 else n,
                bottom = s if r==r2 else n,
            )

# ══════════════════════════════════════════════════════════
# 견적번호 자동 채번
# ══════════════════════════════════════════════════════════
def get_next_quote_number(quote_type: str) -> str:
    """WOCS-YYYY-{TYPE}-NNN 형식으로 자동 채번"""
    year = date.today().year
    prefix = 'A' if quote_type == 'awning' else 'G'
    key = f"{year}_{prefix}"

    data = {}
    if COUNTER_FILE.exists():
        try:
            data = json.loads(COUNTER_FILE.read_text(encoding='utf-8'))
        except:
            data = {}

    seq = data.get(key, 0) + 1
    data[key] = seq
    COUNTER_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                            encoding='utf-8')
    return f"WOCS-{year}-{prefix}{seq:03d}"

def list_quotes():
    """저장된 견적 목록 출력"""
    files = sorted(OUTPUT_DIR.glob('*.xlsx'))
    if not files:
        print("저장된 견적서가 없습니다.")
        return
    print(f"\n{'─'*60}")
    print(f"{'파일명':<40} {'크기':>8}  {'수정일'}")
    print(f"{'─'*60}")
    for f in files:
        stat = f.stat()
        mtime = date.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d')
        size  = f"{stat.st_size/1024:.1f}KB"
        print(f"  {f.name:<38} {size:>8}  {mtime}")
    print(f"{'─'*60}")
    print(f"총 {len(files)}개\n")

# ══════════════════════════════════════════════════════════
# 품목 데이터
# ══════════════════════════════════════════════════════════

GLAMPING_SECTIONS = [
    ('텐트 본체 ①  프레임·조인트·커버', [
        ('WOCS 무용접 유니버설 조인트 (특허)',    'Ø48-60mm',    '세트', '볼트 조립, 용접 불필요'),
        ('아연도금 강관 프레임',                  'Ø48×2.3T',    '본',   '15년 내구'),
        ('방염코팅 PVC 외피 커버',               '650g/m²',     'm²',  'KFI 방염인증'),
        ('방염 아스테이지 — 내부 이너막',        '0.3mm 방염',   'm²',  '이중 단열'),
        ('180° 파노라마 베이윈도 (돔형)',         '강화유리/PVC', '세트', '돔 전용'),
        ('오비띠 — 텐트 고정 보강 벨트',         '50mm 폭',     'm',   '텐트 고정'),
        ('로프·아일렛·카라비너 세트',            '4mm PP',      '세트', ''),
        ('고주파 봉제·재단 제작비',              '',             'm²',  '공장 제작'),
        ('PE 바닥 방수 깔판',                    '1mm 두께',    'm²',  '바닥 방수'),
    ]),
    ('텐트 본체 ②  모델별 본체 (택 1)', [
        ('D-시리즈 돔 5m (D-800)',               'Ø5m / 19.6m²',  '동', '2인 커플형'),
        ('D-시리즈 돔 6m (D-600/800)',           'Ø6m / 28.3m²',  '동', '2~3인'),
        ('D-시리즈 돔 7m (D-600/800)',           'Ø7m / 38.5m²',  '동', '가족형 4인'),
        ('D-시리즈 돔 8m (D-600)',               'Ø8m / 50.3m²',  '동', '로프트 가능'),
        ('D-시리즈 돔 10m',                      'Ø10m / 78.5m²', '동', '리조트 스위트'),
        ('D-시리즈 돔 12m',                      'Ø12m / 113m²',  '동', '대형 이벤트'),
        ('D-시리즈 돔 15m 메가',                 'Ø15m / 177m²',  '동', '2층 로프트 가능'),
        ('S-Classic 사파리 텐트',                '기본형',        '동', '가성비 최상'),
        ('S-Suite 사파리 (우드프레임)',           '원목 프레임',   '동', '숲속 캠프 최적'),
        ('S-Elite 사파리 (프리미엄)',             '32~42m²',      '동', '부티크 리조트'),
        ('S-Extreme 사파리 (극한기후)',           'T5 알루미늄',  '동', '4계절 운영'),
        ('S-Luxury 사파리 빌라',                 '~96m² 멀티룸',  '동', '풀 유리벽'),
        ('Signature-O 코쿤하우스 (오벌)',         '유선형 구조',   '동', '1~2인 독채'),
        ('Signature-A 세일링텐트 (에이펙스)',     '하이 에이펙스',  '동', '4인, 높은 천장'),
        ('Signature-P 버드케이지 (파노라마)',      '360° 파빌리온', '동', '알루미늄 오픈'),
        ('Signature-H 피크로지 (헥사)',           '32~42m² 육각',  '동', '기둥 없는 구조'),
        ('Signature-T 노르딕티피 (삼각)',         '원추형 6인',    '동', '캐노피 연결'),
        ('Signature-Q 큐브캐빈 (모듈)',           '~50m² 큐브',   '동', '모듈러 조립'),
        ('Signature-M 벨텐트 (마이크로)',         '원형 2~4인',    '동', '간편 설치'),
    ]),
    ('데크 및 기초 공사', [
        ('방부목 데크 시공 — SYP 목재',           '38×140mm',    'm²',  '10년 내구 보증'),
        ('데크 하부 기초 블록·합성목',            '',             'm²',  '수평 기초'),
        ('모듈러 데크 (높이조절형)',              '3m×3m 모듈',   'm²',  '200~600mm 조절'),
        ('바닥 방수 처리 — 우레탄 방수',          '2회 도포',     'm²',  '습기 차단'),
        ('배수로 공사',                           '',             'm',   '우수 처리'),
    ]),
    ('전기·설비·부대시설', [
        ('전기 인입 및 분전반 설치',              '30A 이상',     '식',  ''),
        ('내부 LED 조명 설치',                    '웜화이트',     '식',  '무드 조명'),
        ('외부 경관 조명 — 인스타 무드등',        '',             '식',  ''),
        ('냉난방기 설치 (벽걸이형)',              '배관 포함',    '대',  ''),
        ('온수기 설치 (전기식)',                  '',             '대',  '선택'),
        ('콘센트·스위치 설치',                   '',             '식',  ''),
        ('프라이빗 울타리',                       '대나무/우드',  'm',   ''),
        ('모듈러 욕실 (원터치 설치)',             '2~3m 폭',     '식',  '별도 협의'),
        ('태양광 시스템 (3~10kW)',                '독립전원',     '식',  '선택'),
        ('설치부자재 일체 (카라비너 외)',          '',             '식',  ''),
    ]),
    ('애드온·인테리어 패키지 (선택)', [
        ('침실 세트 (침대+테이블+의자)',           '',             '세트', ''),
        ('커튼·파티션 시스템',                    '',             '식',  '룸 분리'),
        ('단열 시스템 (이중단열)',                '-20~40℃',     '식',  '4계절 운영'),
        ('어닝 확장 (입구 캐노피)',               '',             '식',  ''),
        ('바닥재 시스템',                         '',             'm²',  ''),
        ('환기 시스템',                           '',             '식',  ''),
        ('가구·어메니티 패키지',                  '',             '세트', '수건·슬리퍼 등'),
    ]),
    ('시공비', [
        ('현장 답사 및 실측',                    '',             '식',  '시공 전 필수'),
        ('지반 정지 작업',                       '',             'm²',  '수평 작업'),
        ('텐트 설치 인건비 (전국 직시공)',        '',             '식',  'WOCS 직시공 2~3일'),
        ('데크·전기 설치 인건비',                '',             '식',  ''),
        ('인허가 행정 지원 (선택)',               '농지전용 등',  '식',  '별도 협의'),
        ('폐자재 정리·처리비',                   '',             '식',  ''),
        ('운반비 (전국 직시공)',                  '',             '식',  ''),
        ('공과잡비',                             '',             '식',  '소모품 일체'),
    ]),
]


GLAMPING_NOTES = [
    ' ① WOCS 특허 "무용접 다방향 유니버설 조인트" — 내구성·안전성 업계 최상위 등급.',
    ' ② 결제: 계약금 50%(계약시, 입금 완료 후 시공 착수) / 중도금 30%(자재 반입 후 3일 이내) / 잔금 20%(공사 완료 후 7일 이내).',
    ' ③ 하자보증: 시공 완료일로부터 1년 무상 A/S (천재지변·사용자 과실·임의 개조·소모성 부품·자연 열화 제외).',
    ' ④ 지체상금: 공사 지연 시 지체일수 × 계약금액의 1/100 적용 (천재지변·불가항력 제외).',
    ' ⑤ 면책: 본 견적은 개략 견적이며 실측 후 변동 가능 / 지반·장비·인허가 비용 별도 / 건축주 귀책 중단 시 기투입 비용 미반환.',
    ' ⑥ 인허가 지원: 농지전용·관광사업 등록·캠핑장 신고 등 행정 컨설팅 별도 제공.',
    ' ⑦ 세금계산서: 부가세 포함 금액 기준 발행 가능 (요청 시).',
    ' ⑧ 문의: 010-4337-0582 | candlejs6@gmail.com | wocs.kr',
]


# ══════════════════════════════════════════════════════════
# 시트 빌더
# ══════════════════════════════════════════════════════════
def build_sheet(ws, info: dict, sections: list, notes: list):
    """
    info = {
      'quote_no': str,
      'company': str,
      'biz_type': str,
      'item_desc': str,
      'customer': str,
      'address': str,
      'phone': str,
      'work_date': str,
    }
    """
    ws.sheet_view.showGridLines = False
    for col, w in {'A':5,'B':26,'C':12,'D':5,'E':8,'F':13,
                   'G':14,'H':12,'I':14,'J':13}.items():
        ws.column_dimensions[col].width = w

    # ── Row 1: 타이틀 ──
    ws.row_dimensions[1].height = 44
    mc(ws, 'A1:J1')
    sc(ws,'A1','見   積   書',
       Font(name='맑은 고딕',size=26,bold=True,color=C['GTXT']),
       fl(C['DG']), ca(), bd())

    # ── Row 2: 견적번호/날짜 ──
    ws.row_dimensions[2].height = 22
    for rng in ['A2:B2','C2:D2','E2:F2','G2:I2','J2:J2']:
        mc(ws, rng); apply_merge_border(ws, rng)
    sc(ws,'A2','견  적  번  호', fn(9,True,C['W']), fl(C['DG']), ca(), bd())
    sc(ws,'C2', info['quote_no'], fn(10,False,C['DK']), fl(C['CRM']), la(), bd())
    sc(ws,'E2','작  성  일',     fn(9,True,C['W']), fl(C['DG']), ca(), bd())
    sc(ws,'G2','=TEXT(TODAY(),"YYYY년 MM월 DD일")',
       fn(10,True,C['DG']), fl(C['CRM']), ra(), bd())
    sc(ws,'J2','유효: 10일',     fn(9,False,'888888'), fl(C['W']), ca(), bd())

    # ── Rows 3~7: 고객정보 + 공급자 ──
    for r in range(3,8): ws.row_dimensions[r].height = 22
    mc(ws,'F3:F7'); apply_merge_border(ws,'F3:F7')
    sc(ws,'F3','공  급  자', fn(10,True,C['W']), fl(C['DG']), ca(), bd())

    sup_rows = [
        ('사업자\n등록번호', '465-02-03270',      '',       ''),
        ('상  호',          info['company'],      '대표자', '김우성'),
        ('사업장소재지',    '전남 화순군 사평면 유마로 592','',''),
        ('업  태',         info['biz_type'],      '종  목', info['item_desc']),
        ('전  화',         '010-4337-0582',       '사이트', 'wocs.kr'),
    ]
    for i,(l1,v1,l2,v2) in enumerate(sup_rows):
        r = 3+i
        sc(ws,f'G{r}',l1, fn(9,True,C['W']), fl(C['DG']), ca(), bd())
        mc(ws,f'H{r}:I{r}'); apply_merge_border(ws,f'H{r}:I{r}')
        sc(ws,f'H{r}',v1, fn(8 if r==5 else 9), fl(C['W']), la(), bd())
        sc(ws,f'J{r}',f'{l2}: {v2}' if l2 else '',
           fn(9,False,'333333'), fl(C['W']), la(1), bd())

    mc(ws,'A3:E3'); apply_merge_border(ws,'A3:E3')
    sc(ws,'A3','■  수  신  /  고  객  정  보',
       fn(10,True,C['W']), fl(C['DG']), la(), bd())

    cust_data = [
        (4,'고 객 명 (업체명)', info.get('customer','')),
        (5,'현 장 주 소',       info.get('address', '')),
        (6,'연 락 처',          info.get('phone',   '')),
        (7,'시공 희망일',       info.get('work_date','')),
    ]
    for r, lbl, val in cust_data:
        sc(ws,f'A{r}',lbl, fn(9,True,C['W']), fl(C['DG']), ca(), bd())
        mc(ws,f'B{r}:E{r}'); apply_merge_border(ws,f'B{r}:E{r}')
        sc(ws,f'B{r}',val, fn(9), fl(C['CRM']), la(), bd())

    # ── Row 8: 견적 문구 ──
    ws.row_dimensions[8].height = 18
    mc(ws,'A8:J8'); apply_merge_border(ws,'A8:J8')
    sc(ws,'A8','   귀하께서 의뢰하신 건에 대하여 하기와 같이 견적합니다.',
       fn(10,False,'444444'), fl('F6F9F6'), la())

    # ── Row 9: 납품/결재/보증 ──
    ws.row_dimensions[9].height = 22
    for rng in ['A9:B9','C9:D9','E9:F9','G9:H9','I9:J9']:
        mc(ws,rng); apply_merge_border(ws,rng)
    sc(ws,'A9','납품 / 시공일', fn(9,True,C['W']), fl(C['DG']), ca(), bd())
    sc(ws,'C9',info.get('work_date',''), fn(9), fl(C['CRM']), ca(), bd())
    sc(ws,'E9','결  재  방  법', fn(9,True,C['W']), fl(C['DG']), ca(), bd())
    sc(ws,'G9','계약금 50% → 잔금 50%', fn(9), fl(C['W']), ca(), bd())
    sc(ws,'I9','보증기간: 납품 후 1년',  fn(9), fl(C['W']), ca(), bd())

    # ── Row 10: 총금액 ──
    ws.row_dimensions[10].height = 30
    mc(ws,'A10:D10'); mc(ws,'E10:J10')
    apply_merge_border(ws,'A10:D10'); apply_merge_border(ws,'E10:J10')
    sc(ws,'A10','총    금    액    (VAT 포함)',
       Font(name='맑은 고딕',size=13,bold=True,color=C['GTXT']),
       fl(C['DG']), ca(), bd())
    # 총금액 수식은 합계행 확정 후 삽입

    # ── Row 11: 골드 구분선 ──
    ws.row_dimensions[11].height = 4
    mc(ws,'A11:J11'); ws['A11'].fill = fl(C['GOLD'])

    # ── Row 12: 단위 표시 ──
    ws.row_dimensions[12].height = 10
    mc(ws,'H12:J12')
    sc(ws,'H12','(단위 : 원 / 야드 / m / m² / 본 / 개 / 식)',
       fn(8,False,'888888'), align=ra())

    # ── Row 13: 테이블 헤더 ──
    ws.row_dimensions[13].height = 26
    hdrs = [
        ('A','No.',False),('B','품  목  명  /  시  공  내  용',False),
        ('C','규    격',False),('D','단위',False),('E','수  량',False),
        ('F','단  가 (원)',False),('G','공  급  가  액',False),
        ('H','부가세 (10%)',False),('I','합    계',True),('J','비  고',False),
    ]
    for col,hdr,gold in hdrs:
        sc(ws,f'{col}13',hdr,
           fn(10,True,C['W']),
           fl(C['GOLD'] if gold else C['DG']), ca(), bd())

    # ── 품목 행 ──
    cur = 14
    all_rows = []
    item_no = 1

    for sec_title, items in sections:
        ws.row_dimensions[cur].height = 21
        mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
        sc(ws,f'A{cur}',f'  ▶  {sec_title}',
           fn(10,True,C['MDGN']), fl(C['SAGE']), la(2), bd())
        cur += 1

        for name, spec, unit, note in items:
            r = cur
            ws.row_dimensions[r].height = 22
            is_empty = (name == '')
            bg = fl(C['W']) if len(all_rows) % 2 == 0 else fl(C['LGRY'])

            sc(ws,f'A{r}', item_no if not is_empty else '', fn(9), bg, ca(), bd())
            sc(ws,f'B{r}', name, fn(9), bg, la(1), bd())
            sc(ws,f'C{r}', spec, fn(8,False,'555555'), bg, ca(), bd())
            sc(ws,f'D{r}', unit, fn(9), bg, ca(), bd())

            for col_idx, col_letter in [(5,'E'),(6,'F')]:
                c = ws[f'{col_letter}{r}']
                c.fill = fl(C['CRM']) if not is_empty else bg
                c.font = fn(10,True,C['DK'])
                c.alignment = ca() if col_letter == 'E' else ra()
                c.border = bd()
                c.number_format = '#,##0'

            c = ws[f'G{r}']
            c.value = f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r})),E{r}*F{r},"")' if not is_empty else ''
            c.font = fn(10,True); c.fill = bg; c.alignment = ra()
            c.border = bd(); c.number_format = '#,##0'

            c = ws[f'H{r}']
            c.value = f'=IF(ISNUMBER(G{r}),INT(G{r}*0.1),"")' if not is_empty else ''
            c.font = fn(10); c.fill = bg; c.alignment = ra()
            c.border = bd(); c.number_format = '#,##0'

            c = ws[f'I{r}']
            c.value = f'=IF(ISNUMBER(G{r}),G{r}+H{r},"")' if not is_empty else ''
            c.font = Font(name='맑은 고딕',size=10,bold=True,
                          color=C['DG'] if not is_empty else C['DK'])
            c.fill = fl(C['LGOL']) if not is_empty else bg
            c.alignment = ra(); c.border = bd(); c.number_format = '#,##0'

            sc(ws,f'J{r}', note, fn(8,False,'666666'), bg, la(1), bd())

            all_rows.append(r)
            if not is_empty: item_no += 1
            cur += 1

    # 여유 빈 행 3개
    for _ in range(3):
        r = cur; ws.row_dimensions[r].height = 22
        bg = fl(C['W']) if cur % 2 == 0 else fl(C['LGRY'])
        for col in range(1, 11):
            cell = ws.cell(row=r, column=col)
            cell.fill = fl(C['CRM']) if col in (5,6) else bg
            cell.border = bd()
            cell.number_format = '#,##0'
        ws.cell(r,7).value = f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r})),E{r}*F{r},"")'
        ws.cell(r,7).font = fn(10,True); ws.cell(r,7).fill = bg
        ws.cell(r,7).alignment = ra(); ws.cell(r,7).number_format = '#,##0'
        ws.cell(r,8).value = f'=IF(ISNUMBER(G{r}),INT(G{r}*0.1),"")'
        ws.cell(r,8).font = fn(10); ws.cell(r,8).fill = bg
        ws.cell(r,8).alignment = ra(); ws.cell(r,8).number_format = '#,##0'
        ws.cell(r,9).value = f'=IF(ISNUMBER(G{r}),G{r}+H{r},"")'
        ws.cell(r,9).font = Font(name='맑은 고딕',size=10,bold=True,color=C['DG'])
        ws.cell(r,9).fill = fl(C['LGOL']); ws.cell(r,9).alignment = ra()
        ws.cell(r,9).border = bd(); ws.cell(r,9).number_format = '#,##0'
        all_rows.append(r); cur += 1

    fr = all_rows[0]; lr = all_rows[-1]

    # ── 합계 섹션 ──
    ws.row_dimensions[cur].height = 5; cur += 1

    for lbl, formula, is_total in [
        ('공  급  가  액  합  계  (VAT 제외)',
         f'=SUMPRODUCT(IFERROR(G{fr}:G{lr}*1,0))', False),
        ('부  가  세  합  계  (10%)',
         f'=SUMPRODUCT(IFERROR(H{fr}:H{lr}*1,0))', False),
        ('총  합  계  금  액  (VAT 포함)',
         None, True),
    ]:
        ws.row_dimensions[cur].height = 30 if is_total else 24
        mc(ws,f'A{cur}:H{cur}'); apply_merge_border(ws,f'A{cur}:H{cur}')
        sc(ws,f'A{cur}',lbl,
           Font(name='맑은 고딕',size=13 if is_total else 10,bold=True,
                color=C['GTXT'] if is_total else C['W']),
           fl(C['DG'] if is_total else C['GOLD']),
           Alignment(horizontal='right',vertical='center'), bd())
        mc(ws,f'I{cur}:J{cur}'); apply_merge_border(ws,f'I{cur}:J{cur}')
        c = ws[f'I{cur}']
        if is_total:
            c.value = f'=I{cur-2}+I{cur-1}'
        else:
            # v7과 동일하게 배열 수식(Array Formula)으로 저장
            from openpyxl.worksheet.formula import ArrayFormula
            c.value = ArrayFormula(f'I{cur}', formula[1] + formula[1:])  # openpyxl 3.x 버그 우회: 첫글자 중복
        c.font = Font(name='맑은 고딕',size=16 if is_total else 12,bold=True,
                      color=C['DG'] if is_total else C['DK'])
        c.fill = fl(C['LGOL']); c.alignment = ra()
        c.border = bd(); c.number_format = '#,##0'
        if is_total:
            total_row = cur
            ws['E10'].value = f'=I{total_row}'
            ws['E10'].font  = Font(name='맑은 고딕',size=16,bold=True,color=C['DG'])
            ws['E10'].fill  = fl(C['LGOL']); ws['E10'].alignment = ra()
            ws['E10'].border = bd(); ws['E10'].number_format = '#,##0"원 정"'
        cur += 1

    # 한글 금액
    ws.row_dimensions[cur].height = 22
    mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
    c = ws[f'A{cur}']
    c.value = (f'=IF(I{total_row}="","","일금  "&NUMBERSTRING(I{total_row},1)'
               f'&"원정  (부가세 포함)")')
    c.font = Font(name='맑은 고딕',size=12,bold=True,color=C['DG'])
    c.fill = fl(C['SAGE']); c.alignment = ca(); c.border = bd()
    cur += 1

    # ── 특기사항 ──
    ws.row_dimensions[cur].height = 5; cur += 1
    ws.row_dimensions[cur].height = 24
    mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
    sc(ws,f'A{cur}','■  특  기  사  항',
       fn(11,True,C['W']), fl(C['DG']), la(), bd()); cur += 1

    for j, note in enumerate(notes):
        ws.row_dimensions[cur].height = 18
        mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
        sc(ws,f'A{cur}',note, fn(9),
           fl(C['SAGE'] if j%2==0 else C['W']), la(), bd())
        cur += 1

    # ── 서명란 ──
    ws.row_dimensions[cur].height = 5; cur += 1
    ws.row_dimensions[cur].height = 32
    mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
    sc(ws,f'A{cur}','위 견적 내용을 확인하였으며 이에 동의합니다.',
       fn(10), fl(C['W']), ca(), bd()); cur += 1

    ws.row_dimensions[cur].height = 46
    mc(ws,f'A{cur}:B{cur}'); apply_merge_border(ws,f'A{cur}:B{cur}')
    sc(ws,f'A{cur}','공  급  자', fn(10,True,C['W']), fl(C['DG']), ca(), bd())
    mc(ws,f'C{cur}:F{cur}'); apply_merge_border(ws,f'C{cur}:F{cur}')
    sc(ws,f'C{cur}',f"{info['company']}   김우성", fn(10), fl(C['W']), ca(), bd())
    # 도장 이미지 삽입
    import os as _os
    _stamp_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'stamp_clean.png')
    if _os.path.exists(_stamp_path):
        from openpyxl.drawing.image import Image as XlImage
        _stamp = XlImage(_stamp_path)
        _stamp.width = 45
        _stamp.height = 50
        ws.add_image(_stamp, f'F{cur}')
    mc(ws,f'G{cur}:H{cur}'); apply_merge_border(ws,f'G{cur}:H{cur}')
    sc(ws,f'G{cur}','수  급  자', fn(10,True,C['W']), fl(C['DG']), ca(), bd())
    mc(ws,f'I{cur}:J{cur}'); apply_merge_border(ws,f'I{cur}:J{cur}')
    sc(ws,f'I{cur}','                              (인)', fn(10), fl(C['W']), ca(), bd())

    # 인쇄 설정
    ws.print_area = f'A1:J{cur}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0


# ══════════════════════════════════════════════════════════
# 글램핑 텐트 견적서 생성
# ══════════════════════════════════════════════════════════
def generate_glamping_quote(info: dict) -> Path:
    wb = Workbook()
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = '글램핑 텐트 견적서'
    info1 = dict(info)
    info1['company']   = '우성글램핑N텐트'
    info1['biz_type']  = '제조업, 도소매'
    info1['item_desc'] = '텐트·글램핑·캠핑·금속·섬유'
    build_sheet(ws, info1, GLAMPING_SECTIONS, GLAMPING_NOTES)
    today_str = date.today().strftime('%Y%m%d')
    cust_safe = info.get('customer', '').replace(' ','_').replace('/','_') or '견적'
    filename  = f"{today_str}_{info['quote_no']}_{cust_safe}.xlsx"
    out_path  = OUTPUT_DIR / filename
    if wb.calculation is not None:
        wb.calculation.calcId = 0
    wb.save(str(out_path))
    return out_path


def generate_glamping_bytes(info: dict) -> bytes:
    import io as _io
    wb = Workbook()
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = '글램핑 텐트 견적서'
    info1 = dict(info)
    info1['company']   = '우성글램핑N텐트'
    info1['biz_type']  = '제조업, 도소매'
    info1['item_desc'] = '텐트·글램핑·캠핑·금속·섬유'
    build_sheet(ws, info1, GLAMPING_SECTIONS, GLAMPING_NOTES)
    buf = _io.BytesIO()
    if wb.calculation is not None:
        wb.calculation.calcId = 0
    wb.save(buf)
    return buf.getvalue()


if __name__ == '__main__':
    quote_no = get_next_quote_number('glamping')
    print(f"견적번호: {quote_no}")
    customer = input("고객명: ").strip()
    address  = input("현장주소: ").strip()
    phone    = input("연락처: ").strip()
    work_date = input("시공희망일 (Enter=미정): ").strip() or '미정'
    info = {'quote_no': quote_no, 'customer': customer,
            'address': address, 'phone': phone, 'work_date': work_date}
    path = generate_glamping_quote(info)
    print(f"\n완료: {path}")
