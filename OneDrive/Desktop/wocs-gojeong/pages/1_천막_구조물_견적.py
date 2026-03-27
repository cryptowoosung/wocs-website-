import streamlit as st
from PIL import Image, ImageDraw, ImageFont
import os, io, base64, urllib.request
from datetime import datetime
from make_quotation_awning import generate_awning_bytes, get_next_quote_number

def get_font_path():
    f = "NanumGothic.ttf"
    if not os.path.exists(f):
        try:
            urllib.request.urlretrieve(
                "https://github.com/google/fonts/raw/main/ofl/nanumgothic/NanumGothic-Regular.ttf", f)
        except: pass
    return f

FONT_FILE = get_font_path()
LOGO_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "logo.png")

COMPANY_NAME = "우성어닝천막공사캠프시스템"
BIZ_REG_NO = "465-02-03270"
BANK_INFO = "기업은행 323-077581-01-014 (김우성)"
CONTACT_INFO = "010-4337-0582 / wocs.kr"

st.title("WOCS 천막·어닝 구조물 자동 견적 (골조+원단 통합판)")

# ══════════════════════════════════════════════════
# 단가 엔진 — 2025 시장가 기반
# ※ 골조: Soomgo 대형천막 평균 500만/건(250~1,000만), 범위 기준
#   아연도금각파이프 50×50×2.3T 6m=49,000원(m당 ~8,200원)
#   골조 m²단가 = 재료 + 용접·조립 + 도장 + 앵커 + 시공인건비 포함
# ※ 원단: 타포린420 롤 1.2×50m=240,000원(m²당 ~4,000원 원단비)
#   + 재단·고주파·아일렛·설치비 포함 → m²당 총 22,000~55,000원
# ══════════════════════════════════════════════════
def calculate_structure_price(pipe_type, width, length, height):
    """
    골조 단가 = 바닥면적 × m²단가 × 높이보정계수
    ※ 높이 보정 근거:
      - 기둥 파이프 소요량 증가 (4본 기준, +1m = +4m 파이프)
      - 40×40 1.6T 6m=32,000원 → m당 ~5,300원 (11번가 실거래가)
      - 3.5m 이상: 수평 가새(브레이싱) 보강 추가
      - 4.0m 이상: 고소작업 할증 (장비·안전)
    """
    area = width * length
    base_price_per_m2 = {
        "1.4T (경량/기본형)":       210_000,
        "2.0T (표준/강풍대비)":     270_000,
        "3.2T (헤비듀티/적설대비)": 350_000,
    }
    # 높이 보정 계수 (기준 높이 2.5m = 1.0)
    if height <= 2.5:
        h_factor = 1.0
    elif height <= 3.0:
        h_factor = 1.08   # 기둥 연장 + 약간의 보강
    elif height <= 3.5:
        h_factor = 1.18   # 가새 보강 시작
    elif height <= 4.0:
        h_factor = 1.30   # 고소작업 할증
    elif height <= 5.0:
        h_factor = 1.45   # 장비(스카이) 필요 구간
    else:
        h_factor = 1.60   # 대형 구조 + 특수 장비

    return int(area * base_price_per_m2[pipe_type] * h_factor)

def calculate_fabric_price(fabric_type, cover_type, width, length, height):
    if fabric_type == "선택 안 함 (골조만)":
        return 0, 0.0

    fabric_unit_price = {
        "일반 타포린 (방수/기본)":       22_000,  # PVC 방수포 재단+고주파+설치
        "졸타포린 (방염/내구성)":        35_000,  # KFI 방염인증 졸타포린
        "고급 캔버스 (글램핑/어닝전용)": 55_000,  # 고급 PVDF/옥스포드
    }

    roof_area = width * length * 1.2  # 물매 20% 할증
    if cover_type == "지붕만 덮음":
        total_fabric_area = roof_area
    else:
        wall_area = (width + length) * 2 * height
        total_fabric_area = roof_area + wall_area

    return int(total_fabric_area * fabric_unit_price[fabric_type]), total_fabric_area

# ══════════════════════════════════════════════════
# 사이드바 입력
# ══════════════════════════════════════════════════
col1, col2 = st.columns([1, 2])
with col1:
    st.header("1. 현장 정보 및 골조 규격")
    customer_name = st.text_input("고객/현장명", value="고객")

    pipe_type = st.radio("아연도금 각관 두께 (T)", [
        "1.4T (경량/기본형)",
        "2.0T (표준/강풍대비)",
        "3.2T (헤비듀티/적설대비)"
    ])

    col_w, col_l, col_h = st.columns(3)
    with col_w: width_m = st.number_input("가로 (m)", min_value=2.0, max_value=30.0, step=0.5, value=4.0)
    with col_l: length_m = st.number_input("세로 (m)", min_value=2.0, max_value=30.0, step=0.5, value=3.0)
    with col_h: height_m = st.number_input("기둥 높이 (m)", min_value=2.0, max_value=6.0, step=0.5, value=2.5)

    st.markdown("---")
    st.header("2. 천막(원단) 사양")
    fabric_type = st.selectbox("원단 종류", [
        "선택 안 함 (골조만)",
        "일반 타포린 (방수/기본)",
        "졸타포린 (방염/내구성)",
        "고급 캔버스 (글램핑/어닝전용)"
    ])
    cover_type = st.radio("덮개 방식", ["지붕만 덮음", "지붕 + 4면 전체 덮음"], horizontal=True)

    qty = st.number_input("구조물 총 수량", min_value=1, step=1, value=1)

    st.markdown("---")
    st.header("3. 결제 조건")
    col_p1, col_p2, col_p3 = st.columns(3)
    pay1 = col_p1.number_input("선금 %", min_value=0, max_value=100, value=50, step=5)
    pay2 = col_p2.number_input("중도금 %", min_value=0, max_value=100, value=30, step=5)
    pay3 = col_p3.number_input("잔금 %", min_value=0, max_value=100, value=20, step=5)
    if pay1 + pay2 + pay3 != 100:
        st.warning(f"합계 {pay1+pay2+pay3}% — 100%가 되어야 합니다.")

    st.markdown("---")
    custom_note = st.text_area("현장 특이사항 (선택)", placeholder="예: 바닥 콘크리트 별도, 지게차 지원 요망")

# ══════════════════════════════════════════════════
# 계산
# ══════════════════════════════════════════════════
structure_price = calculate_structure_price(pipe_type, width_m, length_m, height_m)
fabric_price, fabric_area = calculate_fabric_price(fabric_type, cover_type, width_m, length_m, height_m)

unit_total = structure_price + fabric_price
total_price = unit_total * qty
vat = int(total_price * 0.1)
final_price = total_price + vat
today_str = datetime.now().strftime("%Y-%m-%d")

pay_text = f"선금 {pay1}% / 중도금 {pay2}% / 잔금 {pay3}% (시공 완료 후)"

# ══════════════════════════════════════════════════
# HTML 화면 출력
# ══════════════════════════════════════════════════
with col2:
    st.header("3. 단가 산출 결과")
    st.markdown(f"**적용 골조 규격:** 가로 {width_m}m x 세로 {length_m}m x 높이 {height_m}m (바닥: {width_m * length_m:.1f}m²)")
    if fabric_area > 0:
        st.markdown(f"**적용 원단 면적:** {fabric_area:.1f} m² ({cover_type})")
    st.markdown(f"**세부 단가(VAT별도):** 골조 {structure_price:,}원 + 원단 {fabric_price:,}원")
    st.markdown(f"**합계 단가(1개당):** {unit_total:,} 원")
    st.markdown(f"**총 결제금액(VAT포함):** <span style='color:red;font-size:24px'>**{final_price:,} 원**</span>", unsafe_allow_html=True)

    # 로고 HTML
    logo_html = ""
    if os.path.exists(LOGO_FILE):
        with open(LOGO_FILE, "rb") as f:
            encoded = base64.b64encode(f.read()).decode()
        logo_html = f'<img src="data:image/png;base64,{encoded}" style="max-height:60px;max-width:160px;margin-right:15px;">'

    stamp_html = """<div style="display:inline-block;border:3px solid red;border-radius:50%;
    width:28px;height:46px;text-align:center;line-height:1.1;color:red;
    font-weight:bold;font-size:12px;margin-left:6px;vertical-align:middle;padding-top:2px;">
    김<br>우<br>성</div>"""

    html = f"""
    <div style="background:#fff;padding:35px;border:1px solid #ddd;box-shadow:0 4px 6px rgba(0,0,0,.1);
    color:#333;font-family:'Malgun Gothic',sans-serif;max-width:800px;margin:auto;">
    <div style="border-bottom:2px solid #333;padding-bottom:15px;margin-bottom:15px;
    display:flex;justify-content:space-between;align-items:center;">
    <div style="display:flex;align-items:center;">
    {logo_html}
    <div><div style="font-size:28px;font-weight:bold;">천막·어닝 구조물 견적서</div></div>
    </div>
    <div style="text-align:right;font-size:13px;line-height:1.5;">
    <strong>{COMPANY_NAME}</strong><br>
    <div style="display:flex;align-items:center;justify-content:flex-end;">
    <span>대표: 김우성</span> {stamp_html}</div>
    | {CONTACT_INFO}<br>사업자번호: {BIZ_REG_NO}<br>
    <span style="color:blue;font-weight:bold;">계좌: {BANK_INFO}</span>
    </div></div>
    <div style="margin-bottom:20px;border-bottom:1px solid #eee;padding-bottom:8px;">
    <strong>수신:</strong> {customer_name} 귀하 <span style="float:right;"><strong>날짜:</strong> {today_str}</span>
    </div>
    <div style="font-size:15px;">
    <div style="display:flex;justify-content:space-between;padding:10px 0;border-bottom:1px solid #eee;">
    <span><strong>품명:</strong> 아연도금 철골구조물 ({pipe_type})</span></div>
    <div style="display:flex;justify-content:space-between;padding:10px 0;border-bottom:1px solid #eee;">
    <span><strong>규격:</strong> {width_m}m(W) × {length_m}m(L) × {height_m}m(H) / 바닥 {width_m*length_m:.1f}m²</span></div>
    """
    if fabric_area > 0:
        html += f"""<div style="display:flex;justify-content:space-between;padding:10px 0;border-bottom:1px solid #eee;">
        <span><strong>원단:</strong> {fabric_type} / {cover_type} (면적 {fabric_area:.1f}m²)</span></div>"""

    html += f"""
    <div style="display:flex;justify-content:space-between;padding:10px 0;border-bottom:1px solid #eee;font-weight:bold;">
    <span>골조 공사비 ({width_m*length_m:.1f}m²)</span><span>{structure_price:,} 원</span></div>
    """
    if fabric_price > 0:
        html += f"""<div style="display:flex;justify-content:space-between;padding:10px 0;border-bottom:1px solid #eee;">
        <span>원단 제작·설치비 ({fabric_area:.1f}m²)</span><span>{fabric_price:,} 원</span></div>"""
    if qty > 1:
        html += f"""<div style="display:flex;justify-content:space-between;padding:10px 0;border-bottom:1px solid #eee;">
        <span>수량: {qty}개 × {unit_total:,}원</span><span>{total_price:,} 원</span></div>"""

    html += f"""
    </div>
    <div style="margin-top:30px;text-align:right;">
    <div style="font-size:15px;color:#555;margin-bottom:4px;">공급가액: {total_price:,} 원</div>
    <div style="font-size:15px;color:#555;margin-bottom:8px;">부가세(VAT): {vat:,} 원</div>
    <div style="font-size:26px;font-weight:bold;color:#d9534f;border-top:2px solid #333;
    padding-top:12px;display:inline-block;">총 견적 금액: {final_price:,} 원</div>
    </div>
    <div style="margin-top:25px;font-size:13px;color:#555;border-top:1px dashed #ccc;padding-top:15px;">
    {'<strong>※ 현장 특이사항:</strong> ' + custom_note + '<br>' if custom_note else ''}
    <hr style="border:none;border-top:1px solid #eee;margin:10px 0;">
    <strong>1. 견적 유효기간:</strong> 견적일로부터 10일 (유효기간 경과 시 재견적 필요)<br>
    <strong style="color:red;">2. 결제 조건:</strong><br>
    <span style="font-size:13px;">
    &nbsp;&nbsp;- <strong>선금 {pay1}%</strong>: 계약 체결 시 납부 (선금 입금 완료 후 시공 착수)<br>
    &nbsp;&nbsp;- <strong>중도금 {pay2}%</strong>: 자재 반입 완료 후 3일 이내 납부<br>
    &nbsp;&nbsp;- <strong>잔금 {pay3}%</strong>: 공사 완료 후 7일 이내 납부<br>
    </span>
    <strong>3. 하자 보증:</strong> 시공 완료일로부터 1년 무상 A/S<br>
    <span style="color:#999;font-size:11px;">&nbsp;&nbsp;(천재지변·사용자 과실·임의 개조·소모성 부품·자연 열화 제외)</span><br>
    <strong>4. 지체상금:</strong> 공사 지연 시 지체일수 × 계약금액의 1/100 적용 (천재지변·불가항력 제외)<br>
    <strong>5. 장비 면책:</strong> 현장 진입 불가 시 소운반 인건비·장비대(스카이차/크레인)는 건축주 별도 부담<br>
    <strong>6. 지반 면책:</strong> 기초 공사 중 발견된 암반·지중 매설물·폐기물 처리 비용은 건축주 책임<br>
    <strong>7. 행정 면책:</strong> 가설건축물 신고·인허가·이웃 민원 해결은 건축주 책임<br>
    <strong>8. 개략 견적:</strong> 본 견적은 현장 실측 전 개략 견적이며, 실측 후 최종 금액이 변동될 수 있습니다.<br>
    <span style="color:#999;font-size:11px;">&nbsp;&nbsp;원자재 가격·환율 변동 시 단가 조정 가능, 건축주 귀책 공사중단 시 기투입 비용 미반환</span><br>
    <strong>9. 세금계산서:</strong> 부가세 포함 금액 기준 발행 가능 (요청 시)
    </div>
    <br>
    <div style="text-align:center;color:#888;font-size:12px;">귀하의 무궁한 발전을 기원합니다.</div>
    </div>
    """

    st.markdown(html, unsafe_allow_html=True)

# ══════════════════════════════════════════════════
# 이미지 생성 + 다운로드
# ══════════════════════════════════════════════════
def create_image():
    img = Image.new('RGB', (800, 1200), color='white')
    draw = ImageDraw.Draw(img)

    try:
        font_L = ImageFont.truetype(FONT_FILE, 28)
        font_M = ImageFont.truetype(FONT_FILE, 20)
        font_S = ImageFont.truetype(FONT_FILE, 15)
        font_XS = ImageFont.truetype(FONT_FILE, 13)
    except:
        font_L = font_M = font_S = font_XS = ImageFont.load_default()

    y = 30

    # 로고
    if os.path.exists(LOGO_FILE):
        try:
            logo = Image.open(LOGO_FILE)
            ar = logo.width / logo.height
            nh = 55; nw = int(nh * ar)
            logo = logo.resize((nw, nh))
            img.paste(logo, (300, y))
            y += 65
        except: pass

    draw.text((230, y), "[ 천막·어닝 구조물 견적서 ]", font=font_L, fill='black'); y += 55

    draw.line((30, y, 770, y), fill='black', width=2); y += 15

    # 고객 정보 (좌측)
    draw.text((30, y), f"고객명: {customer_name} 귀하", font=font_M, fill='black')
    draw.text((30, y + 30), f"발행일: {today_str}", font=font_S, fill='gray')

    # 공급자 (우측)
    draw.text((450, y), "공급자 정보", font=font_M, fill='black'); y += 30
    draw.text((450, y), f"상  호: {COMPANY_NAME}", font=font_S, fill='black'); y += 22
    draw.text((450, y), f"등록번호: {BIZ_REG_NO}", font=font_S, fill='black'); y += 22
    draw.text((450, y), f"연락처: {CONTACT_INFO}", font=font_S, fill='black'); y += 22
    draw.text((450, y), f"계  좌: {BANK_INFO}", font=font_S, fill='blue'); y += 35

    draw.line((30, y, 770, y), fill='gray', width=1); y += 18

    # 품목
    draw.text((30, y), f"품명: 아연도금 철골구조물 ({pipe_type})", font=font_M, fill='black'); y += 28
    draw.text((30, y), f"규격: {width_m}m(W) × {length_m}m(L) × {height_m}m(H) / {width_m*length_m:.1f}m²", font=font_M, fill='black'); y += 28
    if fabric_area > 0:
        draw.text((30, y), f"원단: {fabric_type} / {cover_type} ({fabric_area:.1f}m²)", font=font_M, fill='blue'); y += 28
    draw.text((30, y), f"수량: {qty} 개", font=font_M, fill='black'); y += 35

    draw.line((30, y, 770, y), fill='gray', width=1); y += 18

    # 금액
    draw.text((30, y), f"골조 공사비 ({width_m*length_m:.1f}m²)", font=font_M, fill='black')
    draw.text((750, y), f"{structure_price:,} 원", font=font_M, fill='black', anchor="ra"); y += 28
    if fabric_price > 0:
        draw.text((30, y), f"원단 제작·설치비 ({fabric_area:.1f}m²)", font=font_M, fill='black')
        draw.text((750, y), f"{fabric_price:,} 원", font=font_M, fill='black', anchor="ra"); y += 28

    draw.line((30, y + 8, 770, y + 8), fill='black', width=2); y += 30
    draw.text((450, y), "공급가액:", font=font_M, fill='black')
    draw.text((750, y), f"{total_price:,} 원", font=font_M, fill='black', anchor="ra"); y += 28
    draw.text((450, y), "부 가 세:", font=font_M, fill='black')
    draw.text((750, y), f"{vat:,} 원", font=font_M, fill='black', anchor="ra"); y += 35
    draw.text((450, y), "총 합 계:", font=font_L, fill='red')
    draw.text((750, y), f"{final_price:,} 원", font=font_L, fill='red', anchor="ra"); y += 55

    # 면책조항
    draw.line((30, y, 770, y), fill='black', width=2); y += 12
    draw.text((30, y), "[ WOCS 공사 약관 및 법적 면책 조항 ]", font=font_M, fill='black'); y += 30

    if custom_note:
        draw.text((30, y), f"※ 현장 특이사항: {custom_note}", font=font_S, fill='blue'); y += 22

    clauses = [
        ("1. 유효: 견적일로부터 10일 (경과 시 재견적)", 'gray'),
        (f"2. 결제: 선금{pay1}%(계약시) / 중도금{pay2}%(자재반입후3일)", 'black'),
        (f"   잔금{pay3}%(공사완료후7일내)", 'black'),
        ("3. 하자: 시공완료 후 1년 A/S (과실·개조·소모품 제외)", 'black'),
        ("4. 장비: 진입불가 시 소운반·장비대 건축주 부담", 'red'),
        ("5. 지반: 암반·매설물·폐기물 처리 건축주 책임", 'black'),
        ("6. 행정: 가설건축물 신고·인허가·민원 건축주 책임", 'black'),
        ("7. 귀책: 건축주 사유 공사중단 시 기투입 비용 미반환", 'black'),
        ("8. 본 견적은 개략 견적이며 실측 후 변동 가능", 'gray'),
    ]
    for text, color in clauses:
        draw.text((30, y), text, font=font_XS, fill=color); y += 20

    y += 15
    draw.text((250, y), "귀하의 무궁한 발전을 기원합니다.", font=font_S, fill='gray')

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def create_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "천막구조물 견적서"
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    hdr_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1A3526')
    val_font = Font(name='맑은 고딕', size=10)
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    # 헤더
    ws.merge_cells('A1:D1')
    ws['A1'].value = '천막·어닝 구조물 견적서'
    ws['A1'].font = Font(name='맑은 고딕', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'].value = f'수신: {customer_name} 귀하'
    ws['C2'].value = f'날짜: {today_str}'
    ws['A3'].value = f'공급자: 우성어닝천막공사캠프시스템 (WOCS)'
    ws['C3'].value = f'사업자번호: {BIZ_REG_NO}'

    for col in ['A','B','C','D']:
        ws[f'{col}5'].font = hdr_font
        ws[f'{col}5'].fill = hdr_fill
        ws[f'{col}5'].border = bd
        ws[f'{col}5'].alignment = Alignment(horizontal='center')
    ws['A5'].value = 'No'
    ws['B5'].value = '항목'
    ws['C5'].value = '규격/수량'
    ws['D5'].value = '금액 (원)'

    items = []
    items.append(('골조 공사비', f'{pipe_type} / {width_m}×{length_m}×{height_m}m', structure_price))
    if fabric_price > 0:
        items.append(('원단 제작·설치비', f'{fabric_type} / {fabric_area:.1f}m²', fabric_price))
    if qty > 1:
        items.append(('수량 합산', f'{qty}개', total_price))

    for i, (name, spec, price) in enumerate(items):
        r = 6 + i
        ws[f'A{r}'].value = i + 1
        ws[f'B{r}'].value = name
        ws[f'C{r}'].value = spec
        ws[f'D{r}'].value = price
        ws[f'D{r}'].number_format = '#,##0'
        for col in ['A','B','C','D']:
            ws[f'{col}{r}'].font = val_font
            ws[f'{col}{r}'].border = bd

    r = 6 + len(items) + 1
    for label, val in [('공급가액', total_price), ('부가세(VAT)', vat), ('총 견적 금액', final_price)]:
        ws[f'C{r}'].value = label
        ws[f'C{r}'].font = Font(name='맑은 고딕', size=10, bold=True)
        ws[f'D{r}'].value = val
        ws[f'D{r}'].number_format = '#,##0'
        ws[f'D{r}'].font = Font(name='맑은 고딕', size=12 if '총' in label else 10, bold=True, color='FF0000' if '총' in label else '000000')
        for col in ['C','D']:
            ws[f'{col}{r}'].border = bd
        r += 1

    r += 1
    ws[f'A{r}'].value = f'결제: 선금{pay1}%(계약시) / 중도금{pay2}%(자재반입후3일) / 잔금{pay3}%(공사완료후7일)'
    ws[f'A{r}'].font = Font(name='맑은 고딕', size=9)
    r += 1
    ws[f'A{r}'].value = '하자보증: 시공완료후 1년 / 본 견적은 개략 견적이며 실측 후 변동 가능'
    ws[f'A{r}'].font = Font(name='맑은 고딕', size=9, color='888888')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

st.write("")
col_d1, col_d2, col_d3, col_d4 = st.columns([2, 1, 1, 1])
with col_d2:
    st.download_button(
        "💾 이미지 저장",
        create_image(),
        f"WOCS_천막견적_{customer_name}_{today_str}.png",
        "image/png",
        use_container_width=True
    )
with col_d3:
    st.download_button(
        "📊 Excel 저장",
        create_excel(),
        f"WOCS_천막견적_{customer_name}_{today_str}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
with col_d4:
    quote_no = get_next_quote_number('awning')
    formal_info = {
        'quote_no': quote_no,
        'customer': customer_name,
        'address': '',
        'phone': '',
        'work_date': '미정',
    }
    st.download_button(
        "📋 정식 견적서",
        generate_awning_bytes(formal_info),
        f"WOCS_정식견적_{customer_name}_{today_str}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="시공 확정 시 사용하는 상세 항목별 정식 견적서 (Excel)"
    )
