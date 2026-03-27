import streamlit as st
from PIL import Image, ImageDraw, ImageFont
import os, io, base64, datetime, urllib.request

# ── 폰트 ──
def get_font_path():
    f = "NanumGothic.ttf"
    if not os.path.exists(f):
        try:
            urllib.request.urlretrieve(
                "https://github.com/google/fonts/raw/main/ofl/nanumgothic/NanumGothic-Regular.ttf", f)
        except: pass
    return f

# ── 카탈로그 임포트 ──
from wocs_catalog import CATEGORIES, PRODUCT_CATALOG, ADDON_DEFAULTS
from make_quotation_glamping import generate_glamping_bytes, get_next_quote_number

# ── 기본 설정 ──

MY_BUSINESS_NUM = "465-02-03270"
MY_BANK_INFO = "기업은행 323-077581-01-014 (김우성)"

# ── 로고 ──
logo_path = os.path.join(os.path.dirname(__file__), "logo.png")

# ══════════════════════════════════════════════════
# 사이드바 입력
# ══════════════════════════════════════════════════
with st.sidebar:
    st.title("⛺ WOCS 글램핑 견적")
    st.markdown("---")

    st.markdown("### A. 고객 정보")
    customer_name = st.text_input("고객명 (상호)", value="고객님")

    st.markdown("### B. 제품 선택")
    # 1단계: 카테고리
    cat_key = st.selectbox("카테고리", list(CATEGORIES.keys()),
                           format_func=lambda k: CATEGORIES[k])

    # 2단계: 해당 카테고리 모델
    models = {k: v for k, v in PRODUCT_CATALOG.items() if v['category'] == cat_key}
    model_key = st.selectbox(
        "모델 선택",
        list(models.keys()),
        format_func=lambda k: f"{models[k]['name']}  ({models[k]['size']})"
    )
    product = PRODUCT_CATALOG[model_key]

    # 제품 정보 표시
    st.info(
        f"**{product['name']}**\n\n"
        f"규격: {product['size']} / 면적: {product['area']}m²\n\n"
        f"수용: {product['capacity']}\n\n"
        f"프레임: {product['frame']}\n\n"
        f"납기: {product['lead_time']}"
    )

    qty = st.number_input("수량 (동)", min_value=1, step=1, value=1)

    st.markdown("### C. 부대 옵션")
    include_deck = st.checkbox("데크 시공", value=True)
    deck_area_override = None
    if include_deck:
        deck_area_override = st.number_input(
            "데크 면적 (m²)", min_value=1.0, step=1.0,
            value=float(max(product['area'] * 1.3, 9.0) * qty),
            help="텐트 면적 × 1.3 기본 — 직접 수정 가능")

    include_electric = st.checkbox("전기·조명 설치", value=True)
    include_aircon = st.checkbox("냉난방기", value=False)
    include_hotwater = st.checkbox("온수기", value=False)

    include_fence = st.checkbox("프라이빗 울타리", value=False)
    fence_m = 0
    if include_fence:
        fence_m = st.number_input("울타리 길이 (m)", min_value=1, step=1, value=20)

    include_interior = st.checkbox("인테리어 패키지 (침실세트)", value=False)
    include_insulation = st.checkbox("이중단열 시스템", value=False)

    st.markdown("### D. 추가 비용")
    extra_remove = st.checkbox("기존 시설 철거")
    remove_price = st.number_input("철거비 (원)", value=0 if not extra_remove else 500000,
                                   step=50000, disabled=not extra_remove)
    extra_sky = st.checkbox("장비 사용 (스카이)")
    sky_price = st.number_input("장비비 (원)", value=0 if not extra_sky else 500000,
                                step=50000, disabled=not extra_sky)
    extra_permit = st.checkbox("인허가 행정 지원")
    permit_price = st.number_input("행정 지원비 (원)", value=0 if not extra_permit else 500000,
                                   step=100000, disabled=not extra_permit)

    st.markdown("### E. 결제 조건")
    col_p1, col_p2, col_p3 = st.columns(3)
    pay1 = col_p1.number_input("계약금 %", min_value=0, max_value=100, value=50, step=5)
    pay2 = col_p2.number_input("중도금 %", min_value=0, max_value=100, value=30, step=5)
    pay3 = col_p3.number_input("잔금 %", min_value=0, max_value=100, value=20, step=5)
    if pay1 + pay2 + pay3 != 100:
        st.warning(f"합계 {pay1+pay2+pay3}% — 100%가 되어야 합니다.")

    st.markdown("### F. 기타")
    note_input = st.text_input("비고 (메모)", value="")

# ══════════════════════════════════════════════════
# 자동 계산
# ══════════════════════════════════════════════════
# 본체
body_price = product['ref_price'] * qty

# 데크
deck_price = 0
if include_deck and deck_area_override:
    deck_unit = (ADDON_DEFAULTS['deck_per_m2']
                 + ADDON_DEFAULTS['foundation_per_m2']
                 + ADDON_DEFAULTS['waterproof_per_m2'])
    deck_price = int(deck_area_override * deck_unit)

# 설비
equip_price = 0
equip_items = []
if include_electric:
    equip_price += ADDON_DEFAULTS['electric'] * qty
    equip_items.append(f"전기·조명 ({qty}동)")
if include_aircon:
    equip_price += ADDON_DEFAULTS['aircon'] * qty
    equip_items.append(f"냉난방기 ({qty}대)")
if include_hotwater:
    equip_price += ADDON_DEFAULTS['hotwater'] * qty
    equip_items.append(f"온수기 ({qty}대)")
if include_fence:
    equip_price += ADDON_DEFAULTS['fence_per_m'] * fence_m
    equip_items.append(f"울타리 ({fence_m}m)")
if include_interior:
    equip_price += ADDON_DEFAULTS['interior_set'] * qty
    equip_items.append(f"인테리어 ({qty}세트)")
if include_insulation:
    equip_price += ADDON_DEFAULTS['insulation'] * qty
    equip_items.append(f"이중단열 ({qty}식)")

# 시공비
labor_price = (ADDON_DEFAULTS['labor_install'] * qty
               + (ADDON_DEFAULTS['labor_deck'] if include_deck else 0)
               + ADDON_DEFAULTS['transport']
               + ADDON_DEFAULTS['survey']
               + ADDON_DEFAULTS['misc'])

# 추가비용
extra_price = remove_price + sky_price + permit_price

# 합계
sub_total = body_price + deck_price + equip_price + labor_price + extra_price
vat = int(sub_total * 0.1)
total_price = sub_total + vat
today_str = datetime.datetime.now().strftime("%Y-%m-%d")

# ══════════════════════════════════════════════════
# HTML 견적서 출력
# ══════════════════════════════════════════════════
logo_html = ""
if os.path.exists(logo_path):
    with open(logo_path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode()
    logo_html = f'<img src="data:image/png;base64,{encoded}" style="max-height:80px;max-width:200px;margin-right:20px;">'

stamp_html = """
<div style="display:inline-block;border:3px solid red;border-radius:50%;
width:30px;height:50px;text-align:center;line-height:1.1;color:red;
font-weight:bold;font-size:13px;margin-left:8px;vertical-align:middle;padding-top:2px;">
김<br>우<br>성</div>"""

html = f"""
<div style="background:#fff;padding:40px;border:1px solid #ddd;box-shadow:0 4px 6px rgba(0,0,0,.1);
color:#333;font-family:'Malgun Gothic',sans-serif;max-width:800px;margin:auto;">
<div style="border-bottom:2px solid #333;padding-bottom:20px;margin-bottom:20px;
display:flex;justify-content:space-between;align-items:center;">
<div style="display:flex;align-items:center;">
{logo_html}
<div><div style="font-size:32px;font-weight:bold;">견 적 서</div>
<div style="font-size:14px;color:#666;">글램핑 텐트 · 구조물</div></div>
</div>
<div style="text-align:right;font-size:14px;line-height:1.5;">
<strong>우성어닝천막공사캠프시스템 (WOCS)</strong><br>
<div style="display:flex;align-items:center;justify-content:flex-end;">
<span>대표: 김우성</span> {stamp_html}</div>
| 010-4337-0582<br>사업자번호: {MY_BUSINESS_NUM}<br>
전남 화순군 사평면 유마로 592<br>
<span style="color:blue;font-weight:bold;">계좌: {MY_BANK_INFO}</span>
</div></div>
<div style="margin-bottom:30px;border-bottom:1px solid #eee;padding-bottom:10px;">
<strong>수신:</strong> {customer_name} 귀하 <span style="float:right;"><strong>날짜:</strong> {today_str}</span>
</div>
<div style="font-size:16px;">
"""

def row(label, price, bold=False):
    w = "font-weight:bold;" if bold else ""
    return f'<div style="display:flex;justify-content:space-between;padding:12px 0;border-bottom:1px solid #eee;{w}"><span>{label}</span><span>{price:,} 원</span></div>'

# 본체
html += row(f"⛺ {product['name']} ({product['size']}) × {qty}동", body_price, True)

# 데크
if include_deck and deck_price > 0:
    html += row(f"🪵 데크 시공 ({deck_area_override:.0f}m²)", deck_price)

# 설비
for item_name in equip_items:
    pass  # 개별 표시 대신 합산

if equip_price > 0:
    detail = " / ".join(equip_items)
    html += row(f"⚡ 설비·인테리어 ({detail})", equip_price)

# 시공비
html += row("👷 시공비 (설치+운반+답사+잡비)", labor_price)

# 추가
if remove_price > 0:
    html += row("🏗️ 기존 시설 철거", remove_price)
if sky_price > 0:
    html += row("🚛 장비 사용 (스카이)", sky_price)
if permit_price > 0:
    html += row("📋 인허가 행정 지원", permit_price)

html += f"""
</div>
<div style="margin-top:40px;text-align:right;">
<div style="font-size:16px;color:#555;margin-bottom:5px;">공급가액: {sub_total:,} 원</div>
<div style="font-size:16px;color:#555;margin-bottom:10px;">부가세(VAT): {vat:,} 원</div>
<div style="font-size:28px;font-weight:bold;color:#d9534f;border-top:2px solid #333;
padding-top:15px;display:inline-block;">총 견적 금액: {total_price:,} 원</div>
</div>
<div style="margin-top:30px;font-size:14px;color:#555;border-top:1px dashed #ccc;padding-top:20px;">
{'<strong>※ 특이사항:</strong> ' + note_input + '<br>' if note_input else ''}
<strong>※ 제품 사양:</strong> {product['frame']} / {product['cover']}<br>
<strong>※ 포함 옵션:</strong> {', '.join(product['options'])}<br>
<strong>※ 납기:</strong> {product['lead_time']}<br>
<hr style="border:none;border-top:1px solid #ccc;margin:15px 0;">
<strong>1. 견적 유효기간:</strong> 견적일로부터 10일 (유효기간 경과 시 재견적 필요)<br>
<strong style="color:red;">2. 결제 조건:</strong><br>
<span style="font-size:13px;">
&nbsp;&nbsp;- <strong>계약금 {pay1}%</strong>: 계약 체결 시 납부 (계약금 입금 완료 후 시공 착수)<br>
&nbsp;&nbsp;- <strong>중도금 {pay2}%</strong>: 자재 반입 완료 후 3일 이내 납부<br>
&nbsp;&nbsp;- <strong>잔금 {pay3}%</strong>: 공사 완료 후 7일 이내 납부<br>
</span>
<strong>3. 하자 보증:</strong> 시공 완료일로부터 1년 무상 A/S<br>
<span style="color:#888;font-size:12px;">&nbsp;&nbsp;&nbsp;(천재지변·사용자 과실·임의 개조·소모성 부품·자연 열화 제외)</span><br>
<strong>4. 지체상금:</strong> 공사 지연 시 지체일수 × 계약금액의 1/100 적용 (천재지변·불가항력 제외)<br>
<strong>5. 면책사항:</strong><br>
<span style="color:#888;font-size:12px;">
&nbsp;&nbsp;- 본 견적은 현장 실측 전 <strong>개략 견적</strong>이며, 실측 후 최종 견적이 변동될 수 있습니다.<br>
&nbsp;&nbsp;- 지반 조건·진입로·전기 인입 거리 등 현장 여건에 따라 추가 비용이 발생할 수 있습니다.<br>
&nbsp;&nbsp;- 인허가(농지전용·캠핑장 신고 등) 비용은 별도이며, 행정 결과를 보장하지 않습니다.<br>
&nbsp;&nbsp;- 원자재 가격·환율 변동 시 단가가 조정될 수 있습니다.<br>
&nbsp;&nbsp;- 천재지변·전염병·정부 규제 등 불가항력 사유 시 납기가 변경될 수 있습니다.<br>
&nbsp;&nbsp;- 건축주 귀책사유로 공사가 중단될 경우 기투입 비용은 반환되지 않습니다.<br>
</span>
<strong>6. 세금계산서:</strong> 부가세 포함 금액 기준 발행 가능 (요청 시)<br>
</div>
<br><br>
<div style="text-align:center;color:#888;font-size:13px;">귀하의 무궁한 발전을 기원합니다.</div>
</div>
"""

st.markdown(html, unsafe_allow_html=True)

# ══════════════════════════════════════════════════
# 이미지 저장
# ══════════════════════════════════════════════════
def create_image():
    width, height = 800, 1600
    img = Image.new('RGB', (width, height), color='white')
    draw = ImageDraw.Draw(img)

    font_path = get_font_path()
    try:
        font_L = ImageFont.truetype(font_path, 36)
        font_M = ImageFont.truetype(font_path, 22)
        font_S = ImageFont.truetype(font_path, 18)
        font_B = ImageFont.truetype(font_path, 22)
        font_St = ImageFont.truetype(font_path, 14)
    except:
        font_L = font_M = font_S = font_B = font_St = ImageFont.load_default()

    # 로고
    if os.path.exists(logo_path):
        try:
            li = Image.open(logo_path)
            ar = li.width / li.height
            nh = 80; nw = int(nh * ar)
            li = li.resize((nw, nh))
            img.paste(li, (50, 35))
        except: pass

    draw.text((300, 45), "견  적  서", font=font_L, fill="black")
    draw.text((490, 55), "(글램핑)", font=font_S, fill="#666")
    draw.line((50, 125, 750, 125), fill="black", width=2)

    draw.text((450, 140), "우성어닝천막공사캠프시스템 (WOCS)", font=font_B, fill="black")
    draw.text((450, 175), "대표: 김우성", font=font_S, fill="black")

    # 도장
    sx, sy = 580, 165
    draw.ellipse((sx, sy, sx+30, sy+50), outline="red", width=2)
    draw.text((sx+8, sy+4), "김", font=font_St, fill="red")
    draw.text((sx+8, sy+18), "우", font=font_St, fill="red")
    draw.text((sx+8, sy+32), "성", font=font_St, fill="red")

    ty = 230
    draw.text((450, ty), f"사업자번호: {MY_BUSINESS_NUM}", font=font_S, fill="black")
    draw.text((450, ty+22), "전남 화순군 사평면 유마로 592", font=font_S, fill="black")
    draw.text((450, ty+44), "Tel: 010-4337-0582", font=font_S, fill="black")
    draw.text((450, ty+66), MY_BANK_INFO, font=font_S, fill="blue")

    draw.text((50, 155), f"수신: {customer_name} 귀하", font=font_M, fill="black")
    draw.text((50, 190), f"날짜: {today_str}", font=font_M, fill="black")

    draw.line((50, 340, 750, 340), fill="gray", width=1)
    y = 370

    def draw_row(name, price):
        nonlocal y
        draw.text((50, y), name, font=font_M, fill="black")
        draw.text((750, y), f"{price:,} 원", font=font_M, fill="black", anchor="ra")
        y += 45

    draw_row(f"{product['name']} ({product['size']}) ×{qty}동", body_price)
    if include_deck and deck_price > 0:
        draw_row(f"데크 시공 ({deck_area_override:.0f}m²)", deck_price)
    if equip_price > 0:
        draw_row(f"설비·인테리어", equip_price)
    draw_row("시공비 (설치+운반+답사+잡비)", labor_price)
    if remove_price > 0: draw_row("기존 시설 철거", remove_price)
    if sky_price > 0: draw_row("장비 사용 (스카이)", sky_price)
    if permit_price > 0: draw_row("인허가 행정 지원", permit_price)

    draw.line((50, y+10, 750, y+10), fill="black", width=2)
    y += 40
    draw.text((400, y), "공급가액:", font=font_S, fill="gray")
    draw.text((750, y), f"{sub_total:,} 원", font=font_S, fill="gray", anchor="ra")
    y += 28
    draw.text((400, y), "부가세(VAT):", font=font_S, fill="gray")
    draw.text((750, y), f"{vat:,} 원", font=font_S, fill="gray", anchor="ra")
    y += 45
    draw.text((400, y), "총 견적 금액:", font=font_B, fill="red")
    draw.text((750, y), f"{total_price:,} 원", font=font_B, fill="red", anchor="ra")

    y += 65
    draw.text((50, y), f"제품: {product['frame']}", font=font_S, fill="#555")
    y += 25
    draw.text((50, y), f"옵션: {', '.join(product['options'])}", font=font_S, fill="#555")
    y += 25
    draw.text((50, y), f"납기: {product['lead_time']}", font=font_S, fill="#555")
    y += 35
    if note_input:
        draw.text((50, y), f"※ {note_input}", font=font_S, fill="black")
        y += 30
    draw.text((50, y), "1. 견적 유효기간: 견적일로부터 10일", font=font_S, fill="gray")
    y += 22
    draw.text((50, y), f"2. 결제: 계약금 {pay1}%(계약시) / 중도금 {pay2}%(자재반입후 3일내)", font=font_S, fill="black")
    y += 22
    draw.text((50, y), f"   잔금 {pay3}%(공사완료후 7일내)", font=font_S, fill="black")
    y += 22
    draw.text((50, y), "3. 하자보증: 시공완료일로부터 1년 (천재지변·과실·개조·소모품 제외)", font=font_S, fill="gray")
    y += 22
    draw.text((50, y), "4. 본 견적은 개략 견적이며 실측 후 변동 가능", font=font_S, fill="gray")
    y += 22
    draw.text((50, y), "5. 현장여건·원자재·환율 변동 시 단가 조정 가능", font=font_S, fill="gray")
    y += 22
    draw.text((50, y), "6. 건축주 귀책 공사중단 시 기투입 비용 미반환", font=font_S, fill="gray")
    y += 30
    draw.line((50, y, 750, y), fill="gray", width=1)
    y += 15
    draw.text((50, y), "위 견적 내용을 확인하였으며, 이에 승인하고 계약을 체결합니다.", font=font_S, fill="black")
    y += 35
    draw.text((400, y), "주문 승인 (서명): __________________", font=font_M, fill="black")
    y += 50
    draw.text((250, y), "귀하의 무궁한 발전을 기원합니다.", font=font_S, fill="gray")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def create_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "글램핑텐트 견적서"
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    hdr_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1A3526')
    val_font = Font(name='맑은 고딕', size=10)
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    ws.merge_cells('A1:D1')
    ws['A1'].value = '글램핑 텐트·구조물 견적서'
    ws['A1'].font = Font(name='맑은 고딕', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'].value = f'수신: {customer_name} 귀하'
    ws['C2'].value = f'날짜: {today_str}'
    ws['A3'].value = f'공급자: 우성어닝천막공사캠프시스템 (WOCS)'
    ws['C3'].value = f'사업자번호: {MY_BUSINESS_NUM}'

    for col in ['A','B','C','D']:
        ws[f'{col}5'].font = hdr_font
        ws[f'{col}5'].fill = hdr_fill
        ws[f'{col}5'].border = bd
        ws[f'{col}5'].alignment = Alignment(horizontal='center')
    ws['A5'].value = 'No'
    ws['B5'].value = '항목'
    ws['C5'].value = '규격/수량'
    ws['D5'].value = '금액 (원)'

    items = []
    items.append((f"{product['name']}", f"{product['size']} ×{qty}동", body_price))
    if include_deck and deck_price > 0:
        items.append(('데크 시공', f'{deck_area_override:.0f}m²', deck_price))
    if equip_price > 0:
        items.append(('설비·인테리어', ' / '.join(equip_items), equip_price))
    items.append(('시공비', '설치+운반+답사+잡비', labor_price))
    if remove_price > 0: items.append(('기존 시설 철거', '', remove_price))
    if sky_price > 0: items.append(('장비 사용 (스카이)', '', sky_price))
    if permit_price > 0: items.append(('인허가 행정 지원', '', permit_price))

    for i, (name, spec, price) in enumerate(items):
        r = 6 + i
        ws[f'A{r}'].value = i + 1
        ws[f'B{r}'].value = name
        ws[f'C{r}'].value = spec
        ws[f'D{r}'].value = price
        ws[f'D{r}'].number_format = '#,##0'
        for col in ['A','B','C','D']:
            ws[f'{col}{r}'].font = val_font
            ws[f'{col}{r}'].border = bd

    r = 6 + len(items) + 1
    for label, val in [('공급가액', sub_total), ('부가세(VAT)', vat), ('총 견적 금액', total_price)]:
        ws[f'C{r}'].value = label
        ws[f'C{r}'].font = Font(name='맑은 고딕', size=10, bold=True)
        ws[f'D{r}'].value = val
        ws[f'D{r}'].number_format = '#,##0'
        ws[f'D{r}'].font = Font(name='맑은 고딕', size=12 if '총' in label else 10, bold=True, color='FF0000' if '총' in label else '000000')
        for col in ['C','D']:
            ws[f'{col}{r}'].border = bd
        r += 1

    r += 1
    ws[f'A{r}'].value = f'제품: {product["frame"]} / {product["cover"]}'
    ws[f'A{r}'].font = Font(name='맑은 고딕', size=9)
    r += 1
    ws[f'A{r}'].value = f'결제: 계약금{pay1}%(계약시) / 중도금{pay2}%(자재반입후3일) / 잔금{pay3}%(공사완료후7일)'
    ws[f'A{r}'].font = Font(name='맑은 고딕', size=9)
    r += 1
    ws[f'A{r}'].value = '하자보증: 시공완료후 1년 / 본 견적은 개략 견적이며 실측 후 변동 가능'
    ws[f'A{r}'].font = Font(name='맑은 고딕', size=9, color='888888')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

st.write("")
st.write("")
col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
with col2:
    st.download_button("💾 이미지 저장", create_image(),
                       f"글램핑견적_{customer_name}_{today_str}.png",
                       "image/png", use_container_width=True)
with col3:
    st.download_button("📊 Excel 저장", create_excel(),
                       f"글램핑견적_{customer_name}_{today_str}.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)
with col4:
    quote_no = get_next_quote_number('glamping')
    formal_info = {
        'quote_no': quote_no,
        'customer': customer_name,
        'address': '',
        'phone': '',
        'work_date': '미정',
    }
    st.download_button(
        "📋 정식 견적서",
        generate_glamping_bytes(formal_info),
        f"WOCS_정식견적_{customer_name}_{today_str}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="시공 확정 시 사용하는 상세 항목별 정식 견적서 (Excel)"
    )
