import streamlit as st
import datetime
import io
import base64
import os
import urllib.request
from make_quotation_awning import get_next_quote_number

# 라이브러리 체크
try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    st.error("Pillow가 설치되지 않았습니다. 터미널에 'pip install Pillow'를 입력해주세요.")
    st.stop()

# -----------------------------------------------------------------------------
# 0. 폰트 자동 설치
# -----------------------------------------------------------------------------
def get_font_path():
    font_filename = "NanumGothic.ttf"
    if not os.path.exists(font_filename):
        url = "https://github.com/google/fonts/raw/main/ofl/nanumgothic/NanumGothic-Regular.ttf"
        try:
            urllib.request.urlretrieve(url, font_filename)
        except:
            pass
    return font_filename

# -----------------------------------------------------------------------------
# 1. 기본 설정
# -----------------------------------------------------------------------------

# 사장님 정보
MY_BUSINESS_NUM = "465-02-03270"        
MY_BANK_INFO = "기업은행 323-077581-01-014 (김우성)" 

# -----------------------------------------------------------------------------
# 2. 데이터 (단가표 DB)
# -----------------------------------------------------------------------------

# 1) 다각형 (최소 2.4m)
polygon_prices = {
    "600 x 600": 96000,
    "600 x 900": 102000,
    "900 x 900": 108000,
    "1200 x 900": 114000,
    "1200 x 1200": 120000,
    "1200 x 1500": 144000,
    "1500 x 1500": 156000,
    "1800 x 1800": 192000,
    "2100 x 2100": 228000
}

# 2) 라운드형 (최소 기준 다름)
round_prices = {
    "600 x 600": 144000,
    "900 x 900": 168000,
    "1200 x 1200": 192000,
    "1500 x 1500": 216000
}

# -----------------------------------------------------------------------------
# 3. 사이드바 입력
# -----------------------------------------------------------------------------
with st.sidebar:
    st.title("🏛️ 고정식 어닝 견적")

    st.markdown("---")
    
    st.markdown("### A. 형태 및 규격")
    customer_name = st.text_input("고객명 (상호)", value="고객님")
    
    # 1. 어닝 대분류 선택
    category = st.radio("어닝 종류", ["다각형 (박스/코너)", "라운드 (바가지/바나나)"], horizontal=True)
    
    # 2. 세부 형태 및 규격 설정
    if category == "다각형 (박스/코너)":
        shape_type = st.selectbox("상세 형태", ["다각형", "사각(박스형)", "코너(라운드)"])
        selected_spec = st.selectbox("규격 (높이 x 돌출)", list(polygon_prices.keys()))
        unit_price = polygon_prices[selected_spec]
        min_width = 2.4 # 다각형 기본 최소길이
        
    else: # 라운드형
        shape_type = st.selectbox("상세 형태", ["바가지형", "옆바가지형", "반 바나나형", "좌우 바나나형"])
        selected_spec = st.selectbox("규격 (높이 x 돌출)", list(round_prices.keys()))
        unit_price = round_prices[selected_spec]
        
        # 형태별 최소 기준 적용
        if shape_type in ["바가지형", "반 바나나형"]:
            min_width = 2.0
        else: # 옆바가지형, 좌우 바나나형
            min_width = 2.4

    # 3. 가로 길이 입력
    width_input = st.number_input("가로 길이 (m)", min_value=0.5, step=0.1, value=2.0)
    
    calc_width = max(width_input, min_width)
    if width_input < min_width:
        st.info(f"💡 선택하신 '{shape_type}'은 최소 {min_width}m 기준으로 계산됩니다.")

    st.markdown("### B. 원단 및 조명")
    fabric_type = st.radio("원단 종류", ["국산 (방수)", "수입 (어닝전용)"], horizontal=True)
    fabric_add_price = st.number_input("원단 추가금 (원)", value=0, step=10000)
    
    use_light = st.checkbox("T5 조명 설치")
    light_count = st.number_input("조명 개수 (개)", value=0 if not use_light else 1, step=1, disabled=not use_light)
    light_price = 20000 * light_count 

    st.markdown("### C. 기본 옵션")
    use_print = st.checkbox("로고 인쇄 (1도/1면)")
    print_price = st.number_input("인쇄비 (원)", value=0 if not use_print else 100000, step=10000, disabled=not use_print)
    
    use_gutter = st.checkbox("물받이 설치")
    default_gutter_price = int(width_input * 10000)
    gutter_price = st.number_input("물받이 가격 (원)", value=0 if not use_gutter else default_gutter_price, step=5000, disabled=not use_gutter)

    st.markdown("### D. 시공 및 부자재")
    labor_price = st.number_input("시공 노무비 (원)", value=250000, step=10000)
    material_price = st.number_input("부자재비용 (원)", value=100000, step=5000)
    transport_price = st.number_input("운송비용 (원)", value=60000, step=5000)

    st.markdown("---")
    st.markdown("### E. 현장 특수 조건")
    
    use_remove = st.checkbox("기존 철거")
    remove_price = st.number_input("철거비용 (원)", value=0 if not use_remove else 150000, step=10000, disabled=not use_remove)

    use_ladder = st.checkbox("장비 사용 (스카이)")
    ladder_price = st.number_input("장비 사용료 (원)", value=0 if not use_ladder else 500000, step=10000, disabled=not use_ladder)

    use_bracket = st.checkbox("특수 브라켓 추가")
    bracket_count = st.number_input("브라켓 개수 (개)", value=0 if not use_bracket else 1, step=1, disabled=not use_bracket)
    bracket_price = 20000 * bracket_count 

    use_special_wall = st.checkbox("특수 벽면 시공 (대리석/유리)")
    special_wall_price = st.number_input("시공비 추가 (원)", value=0 if not use_special_wall else 80000, step=10000, disabled=not use_special_wall)

    use_pole = st.checkbox("보조 기둥 설치")
    pole_price = st.number_input("기둥 설치비 (원)", value=0 if not use_pole else 100000, step=10000, disabled=not use_pole)

    use_night = st.checkbox("야간/주말 할증")
    night_price = st.number_input("할증 비용 (원)", value=0 if not use_night else 50000, step=10000, disabled=not use_night)

    st.markdown("### F. 기타/특이사항")
    note_input = st.text_input("비고 (메모)", value="")

# -----------------------------------------------------------------------------
# 4. 계산 로직
# -----------------------------------------------------------------------------
base_price = int(calc_width * unit_price)

sub_total = (base_price + fabric_add_price + light_price + print_price + gutter_price +
             labor_price + material_price + transport_price +
             remove_price + ladder_price + bracket_price + special_wall_price + pole_price + night_price)

vat = int(sub_total * 0.1)
total_price = sub_total + vat
today_str = datetime.datetime.now().strftime("%Y-%m-%d")

spec_info_text = f"{shape_type} ({selected_spec})"
if width_input < min_width:
    spec_info_text += f" / {width_input}m (최소 {min_width}m 적용)"
else:
    spec_info_text += f" / {width_input}m"

# -----------------------------------------------------------------------------
# 5. HTML 화면 출력
# -----------------------------------------------------------------------------
logo_html = ""
logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
if os.path.exists(logo_path):
    with open(logo_path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode()
    logo_html = f'<img src="data:image/png;base64,{encoded}" style="max-height: 80px; max-width: 200px; margin-right: 20px;">'

stamp_html = """
<div style="
    display: inline-block;
    border: 3px solid red;
    border-radius: 50%;
    width: 30px;
    height: 50px;
    text-align: center;
    line-height: 1.1;
    color: red;
    font-weight: bold;
    font-size: 13px;
    margin-left: 8px;
    vertical-align: middle;
    padding-top: 2px;
">
    김<br>우<br>성
</div>
"""

html_content = f"""
<div style="background-color: white; padding: 40px; border: 1px solid #ddd; box-shadow: 0 4px 6px rgba(0,0,0,0.1); color: #333; font-family: 'Malgun Gothic', sans-serif; max-width: 800px; margin: auto;">
<div style="border-bottom: 2px solid #333; padding-bottom: 20px; margin-bottom: 20px; display: flex; justify-content: space-between; align-items: center;">
<div style="display: flex; align-items: center;">
{logo_html}
<div style="font-size: 32px; font-weight: bold;">견 적 서 <span style="font-size:18px; color:#666;">(고정식)</span></div>
</div>
<div style="text-align: right; font-size: 14px; line-height: 1.5;">
<strong>우성어닝천막공사캠프시스템 (WOCS)</strong><br>
<div style="display: flex; align-items: center; justify-content: flex-end;">
    <span>대표: 김우성</span> {stamp_html}
</div>
| 010-4337-0582<br>
사업자번호: {MY_BUSINESS_NUM}<br>
전남 화순군 사평면 유마로 592<br>
<span style="color: blue; font-weight: bold;">계좌: {MY_BANK_INFO}</span>
</div>
</div>
<div style="margin-bottom: 30px; border-bottom: 1px solid #eee; padding-bottom: 10px;">
<strong>수신:</strong> {customer_name} 귀하 <span style="float:right;"><strong>날짜:</strong> {today_str}</span>
</div>
<div style="font-size: 16px;">
<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;">
<span>🏷️ <strong>{spec_info_text}</strong></span>
<span style="font-weight:bold;">{base_price:,} 원</span>
</div>
"""

if fabric_add_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🧵 원단 추가 ({fabric_type})</span><span>+{fabric_add_price:,} 원</span></div>"""
if use_print and print_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🎨 로고 인쇄</span><span>+{print_price:,} 원</span></div>"""
if use_light and light_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>💡 T5 조명 ({light_count}개)</span><span>+{light_price:,} 원</span></div>"""
if use_gutter and gutter_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>💧 물받이 설치</span><span>+{gutter_price:,} 원</span></div>"""

if use_remove and remove_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🏗️ 기존 철거</span><span>+{remove_price:,} 원</span></div>"""
if use_ladder and ladder_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🚛 장비 사용 (스카이)</span><span>+{ladder_price:,} 원</span></div>"""
if use_bracket and bracket_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🔧 특수 브라켓 ({bracket_count}개)</span><span>+{bracket_price:,} 원</span></div>"""
if use_special_wall and special_wall_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🧱 특수 벽면 시공</span><span>+{special_wall_price:,} 원</span></div>"""
if use_pole and pole_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🏛️ 보조 기둥 설치</span><span>+{pole_price:,} 원</span></div>"""
if use_night and night_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🌙 야간/주말 할증</span><span>+{night_price:,} 원</span></div>"""

html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>👷 시공 노무비</span><span>+{labor_price:,} 원</span></div>"""
if material_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🔩 부자재비용</span><span>+{material_price:,} 원</span></div>"""
if transport_price > 0:
    html_content += f"""<div style="display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px solid #eee;"><span>🚚 운송비용</span><span>+{transport_price:,} 원</span></div>"""

html_content += f"""
</div>
<div style="margin-top: 40px; text-align: right;">
<div style="font-size: 16px; color: #555; margin-bottom: 5px;">공급가액: {sub_total:,} 원</div>
<div style="font-size: 16px; color: #555; margin-bottom: 10px;">부가세(VAT): {vat:,} 원</div>
<div style="font-size: 28px; font-weight: bold; color: #d9534f; border-top: 2px solid #333; padding-top: 15px; display: inline-block;">총 견적 금액: {total_price:,} 원</div>
</div>
<div style="margin-top: 30px; font-size: 14px; color: #555; border-top: 1px dashed #ccc; padding-top: 20px;">
{'<strong>※ 특이사항:</strong> ' + note_input + '<br>' if note_input else ''}
<strong>1. 견적 유효기간:</strong> 견적일로부터 10일 (유효기간 경과 시 재견적 필요)<br>
<strong style="color:red;">2. 결제 조건:</strong><br>
<span style="font-size:13px;">
&nbsp;&nbsp;- <strong>계약금 50%</strong>: 계약 체결 시 납부 (입금 완료 후 시공 착수)<br>
&nbsp;&nbsp;- <strong>중도금 30%</strong>: 자재 반입 완료 후 3일 이내 납부<br>
&nbsp;&nbsp;- <strong>잔금 20%</strong>: 공사 완료 후 7일 이내 납부<br>
</span>
<strong>3. 하자 보증:</strong> 시공 완료일로부터 1년 무상 A/S (천재지변·사용자 과실·임의 개조·소모성 부품·자연 열화 제외)<br>
<strong>4. 지체상금:</strong> 공사 지연 시 지체일수 × 계약금액의 1/100 적용 (천재지변·불가항력 제외)<br>
<strong>5. 면책:</strong> 본 견적은 개략 견적이며 실측 후 변동 가능 / 장비·지반·인허가 비용 별도 / 건축주 귀책 중단 시 기투입 비용 미반환<br>
<strong>6. 세금계산서:</strong> 부가세 포함 금액 기준 발행 가능 (요청 시)
</div>
<br><br>
<div style="text-align:center; color:#888; font-size:13px;">귀하의 무궁한 발전을 기원합니다.</div>
</div>
"""

st.markdown(html_content, unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# 6. 이미지 저장
# -----------------------------------------------------------------------------
def create_image():
    width, height = 800, 1450
    img = Image.new('RGB', (width, height), color='white')
    draw = ImageDraw.Draw(img)
    
    font_path = get_font_path()
    try:
        font_L = ImageFont.truetype(font_path, 40)
        font_M = ImageFont.truetype(font_path, 25)
        font_S = ImageFont.truetype(font_path, 20)
        font_Bold = ImageFont.truetype(font_path, 25)
        font_Stamp = ImageFont.truetype(font_path, 14)
    except:
        font_L = ImageFont.load_default()
        font_M = ImageFont.load_default()
        font_S = ImageFont.load_default()
        font_Bold = ImageFont.load_default()
        font_Stamp = ImageFont.load_default()

    if os.path.exists(logo_path):
        try:
            logo_img = Image.open(logo_path)
            aspect_ratio = logo_img.width / logo_img.height
            new_height = 80
            new_width = int(new_height * aspect_ratio)
            logo_img = logo_img.resize((new_width, new_height))
            img.paste(logo_img, (50, 40))
        except:
            pass

    draw.text((320, 50), "견  적  서", font=font_L, fill="black")
    draw.text((500, 65), "(고정식)", font=font_S, fill="#666")

    draw.line((50, 130, 750, 130), fill="black", width=2)
    
    draw.text((450, 150), "우성어닝천막공사캠프시스템 (WOCS)", font=font_Bold, fill="black")
    draw.text((450, 190), "대표: 김우성", font=font_S, fill="black")

    # 도장
    stamp_x = 580
    stamp_y = 178
    stamp_w = 30
    stamp_h = 50
    draw.ellipse((stamp_x, stamp_y, stamp_x + stamp_w, stamp_y + stamp_h), outline="red", width=2)
    draw.text((stamp_x + 8, stamp_y + 4), "김", font=font_Stamp, fill="red")
    draw.text((stamp_x + 8, stamp_y + 18), "우", font=font_Stamp, fill="red")
    draw.text((stamp_x + 8, stamp_y + 32), "성", font=font_Stamp, fill="red")

    # 정보
    text_start_y = 245
    draw.text((450, text_start_y), f"사업자번호: {MY_BUSINESS_NUM}", font=font_S, fill="black")
    draw.text((450, text_start_y + 25), "전남 화순군 사평면 유마로 592", font=font_S, fill="black")
    draw.text((450, text_start_y + 50), "Tel: 010-4337-0582", font=font_S, fill="black")
    draw.text((450, text_start_y + 75), f"{MY_BANK_INFO}", font=font_S, fill="blue")

    draw.text((50, 170), f"수신: {customer_name} 귀하", font=font_M, fill="black")
    draw.text((50, 210), f"날짜: {today_str}", font=font_M, fill="black")

    line_y = 380 
    draw.line((50, line_y, 750, line_y), fill="gray", width=1)
    y = line_y + 30
    def draw_row(name, price):
        nonlocal y
        draw.text((50, y), name, font=font_M, fill="black")
        draw.text((750, y), f"{price:,} 원", font=font_M, fill="black", anchor="ra")
        y += 50

    draw_row(f"{spec_info_text}", base_price)
    
    if fabric_add_price > 0: draw_row(f"원단 추가 ({fabric_type})", fabric_add_price)
    if use_print and print_price > 0: draw_row("로고 인쇄", print_price)
    if use_light and light_price > 0: draw_row(f"T5 조명 ({light_count}개)", light_price)
    if use_gutter and gutter_price > 0: draw_row("물받이 설치", gutter_price)
    
    if use_remove and remove_price > 0: draw_row("기존 철거", remove_price)
    if use_ladder and ladder_price > 0: draw_row("장비 사용 (스카이)", ladder_price)
    if use_bracket and bracket_price > 0: draw_row(f"특수 브라켓 ({bracket_count}개)", bracket_price)
    if use_special_wall and special_wall_price > 0: draw_row("특수 벽면 시공", special_wall_price)
    if use_pole and pole_price > 0: draw_row("보조 기둥 설치", pole_price)
    if use_night and night_price > 0: draw_row("야간/주말 할증", night_price)
    
    draw_row("시공 노무비", labor_price)
    if material_price > 0: draw_row("부자재비용", material_price)
    if transport_price > 0: draw_row("운송비용", transport_price)

    draw.line((50, y+10, 750, y+10), fill="black", width=2)
    y += 40
    draw.text((400, y), "공급가액:", font=font_S, fill="gray")
    draw.text((750, y), f"{sub_total:,} 원", font=font_S, fill="gray", anchor="ra")
    y += 30
    draw.text((400, y), "부가세(VAT):", font=font_S, fill="gray")
    draw.text((750, y), f"{vat:,} 원", font=font_S, fill="gray", anchor="ra")
    y += 50
    draw.text((400, y), "총 견적 금액:", font=font_Bold, fill="red")
    draw.text((750, y), f"{total_price:,} 원", font=font_Bold, fill="red", anchor="ra")
    
    y += 70
    if note_input:
        draw.text((50, y), f"※ 특이사항: {note_input}", font=font_S, fill="black")
        y += 40
    draw.text((50, y), "1. 견적 유효기간: 견적일로부터 10일", font=font_S, fill="gray")
    y += 25
    draw.text((50, y), "2. 결제: 계약금50%(계약시) / 중도금30%(자재반입후3일)", font=font_S, fill="black")
    y += 25
    draw.text((50, y), "   잔금20%(공사완료후7일내)", font=font_S, fill="black")
    y += 25
    draw.text((50, y), "3. 하자보증: 시공완료후 1년 (과실·개조·소모품·자연열화 제외)", font=font_S, fill="gray")
    y += 25
    draw.text((50, y), "4. 지체상금: 지체일수 × 계약금액의 1/100 (불가항력 제외)", font=font_S, fill="black")
    y += 25
    draw.text((50, y), "5. 개략견적이며 실측후 변동가능 / 귀책중단시 비용미반환", font=font_S, fill="gray")
    
    y += 50
    draw.line((50, y, 750, y), fill="gray", width=1)
    y += 20
    draw.text((50, y), "위 견적 내용을 확인하였으며, 이에 승인하고 계약을 체결합니다.", font=font_S, fill="black")
    y += 40
    draw.text((400, y), "주문 승인 (서명): __________________", font=font_M, fill="black")

    y += 60
    draw.text((250, y), "귀하의 무궁한 발전을 기원합니다.", font=font_S, fill="gray")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def create_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "고정식어닝 견적서"
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    hdr_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1A3526')
    val_font = Font(name='맑은 고딕', size=10)
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    ws.merge_cells('A1:D1')
    ws['A1'].value = '견적서 (고정식 어닝)'
    ws['A1'].font = Font(name='맑은 고딕', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].value = f'수신: {customer_name} 귀하'
    ws['C2'].value = f'날짜: {today_str}'
    ws['A3'].value = '공급자: 우성어닝천막공사캠프시스템 (WOCS)'
    ws['C3'].value = f'사업자번호: {MY_BUSINESS_NUM}'

    for col in ['A','B','C','D']:
        ws[f'{col}5'].font = hdr_font
        ws[f'{col}5'].fill = hdr_fill
        ws[f'{col}5'].border = bd
        ws[f'{col}5'].alignment = Alignment(horizontal='center')
    ws['A5'].value = 'No'
    ws['B5'].value = '항목'
    ws['C5'].value = '규격/수량'
    ws['D5'].value = '금액 (원)'

    items = [(spec_info_text, f'가로 {width_input}m', base_price)]
    if fabric_add_price > 0: items.append((f'원단 추가 ({fabric_type})', '', fabric_add_price))
    if use_print and print_price > 0: items.append(('로고 인쇄', '1도/1면', print_price))
    if use_light and light_price > 0: items.append((f'T5 조명', f'{light_count}개', light_price))
    if use_gutter and gutter_price > 0: items.append(('물받이 설치', '', gutter_price))
    if use_remove and remove_price > 0: items.append(('기존 철거', '', remove_price))
    if use_ladder and ladder_price > 0: items.append(('장비 사용 (스카이)', '', ladder_price))
    if use_bracket and bracket_price > 0: items.append((f'특수 브라켓', f'{bracket_count}개', bracket_price))
    if use_special_wall and special_wall_price > 0: items.append(('특수 벽면 시공', '', special_wall_price))
    if use_pole and pole_price > 0: items.append(('보조 기둥 설치', '', pole_price))
    if use_night and night_price > 0: items.append(('야간/주말 할증', '', night_price))
    items.append(('시공 노무비', '', labor_price))
    if material_price > 0: items.append(('부자재비용', '', material_price))
    if transport_price > 0: items.append(('운송비용', '', transport_price))

    for i, (name, spec, price) in enumerate(items):
        r = 6 + i
        ws[f'A{r}'].value = i + 1
        ws[f'B{r}'].value = name
        ws[f'C{r}'].value = spec
        ws[f'D{r}'].value = price
        ws[f'D{r}'].number_format = '#,##0'
        for col in ['A','B','C','D']:
            ws[f'{col}{r}'].font = val_font
            ws[f'{col}{r}'].border = bd

    r = 6 + len(items) + 1
    for label, val in [('공급가액', sub_total), ('부가세(VAT)', vat), ('총 견적 금액', total_price)]:
        ws[f'C{r}'].value = label
        ws[f'C{r}'].font = Font(name='맑은 고딕', size=10, bold=True)
        ws[f'D{r}'].value = val
        ws[f'D{r}'].number_format = '#,##0'
        ws[f'D{r}'].font = Font(name='맑은 고딕', size=12 if '총' in label else 10, bold=True, color='FF0000' if '총' in label else '000000')
        for col in ['C','D']: ws[f'{col}{r}'].border = bd
        r += 1
    r += 1
    ws[f'A{r}'].value = '결제: 계약금50%(계약시) / 중도금30%(자재반입후3일) / 잔금20%(공사완료후7일)'
    ws[f'A{r}'].font = Font(name='맑은 고딕', size=9)
    r += 1
    ws[f'A{r}'].value = '하자보증: 시공완료후 1년 / 본 견적은 개략 견적이며 실측 후 변동 가능'
    ws[f'A{r}'].font = Font(name='맑은 고딕', size=9, color='888888')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def create_formal_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.formula import ArrayFormula
    wb = Workbook()
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = "고정식 어닝 정식 견적서"
    ws.sheet_view.showGridLines = False

    DG = '1A3526'; GOLD = 'B59035'; SAGE = 'E8F0EA'; CRM = 'FDF8EE'; LGRY = 'F4F6F4'; LGOL = 'FDF3DC'
    def fl(c): return PatternFill('solid', fgColor=c)
    def fn(sz=10, bold=False, color='1A1A1A'): return Font(name='맑은 고딕', size=sz, bold=bold, color=color)
    def bd():
        s = Side('thin')
        return Border(left=s, right=s, top=s, bottom=s)

    for col, w in {'A':5,'B':26,'C':12,'D':5,'E':8,'F':13,'G':14,'H':12,'I':14,'J':13}.items():
        ws.column_dimensions[col].width = w

    # Row 1: 타이틀
    ws.merge_cells('A1:J1')
    ws['A1'].value = '見   積   書  (고정식 어닝)'
    ws['A1'].font = Font(name='맑은 고딕', size=22, bold=True, color='F0D080')
    ws['A1'].fill = fl(DG); ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Row 2: 견적번호/날짜
    quote_no = get_next_quote_number('awning')
    ws.merge_cells('A2:B2'); ws['A2'].value = '견적번호'; ws['A2'].font = fn(9,True,'FFFFFF'); ws['A2'].fill = fl(DG); ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C2:D2'); ws['C2'].value = quote_no; ws['C2'].font = fn(10); ws['C2'].fill = fl(CRM)
    ws.merge_cells('E2:F2'); ws['E2'].value = '작성일'; ws['E2'].font = fn(9,True,'FFFFFF'); ws['E2'].fill = fl(DG); ws['E2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('G2:I2'); ws['G2'].value = '=TEXT(TODAY(),"YYYY년 MM월 DD일")'; ws['G2'].font = fn(10,True,DG); ws['G2'].fill = fl(CRM)
    ws['J2'].value = '유효:10일'; ws['J2'].font = fn(9,False,'888888')

    # Row 3~7: 고객+공급자
    ws.merge_cells('A3:E3'); ws['A3'].value = '■ 수신 / 고객 정보'; ws['A3'].font = fn(10,True,'FFFFFF'); ws['A3'].fill = fl(DG)
    ws.merge_cells('F3:F7'); ws['F3'].value = '공급자'; ws['F3'].font = fn(10,True,'FFFFFF'); ws['F3'].fill = fl(DG); ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
    for r, lbl, val, rl, rv in [
        (3,'','','사업자번호','465-02-03270'),
        (4,'고객명',customer_name,'상호','우성어닝천막공사캠프시스템'),
        (5,'현장주소','','소재지','전남 화순군 사평면 유마로 592'),
        (6,'연락처','','업태/종목','제조업/어닝·천막·방수포'),
        (7,'시공희망일','','전화/사이트','010-4337-0582 / wocs.kr'),
    ]:
        if lbl:
            ws[f'A{r}'].value = lbl; ws[f'A{r}'].font = fn(9,True,'FFFFFF'); ws[f'A{r}'].fill = fl(DG); ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f'B{r}:E{r}'); ws[f'B{r}'].value = val; ws[f'B{r}'].fill = fl(CRM)
        ws[f'G{r}'].value = rl; ws[f'G{r}'].font = fn(9,True,'FFFFFF'); ws[f'G{r}'].fill = fl(DG); ws[f'G{r}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'H{r}:J{r}'); ws[f'H{r}'].value = rv; ws[f'H{r}'].font = fn(9)

    # Row 8: 안내문
    ws.merge_cells('A8:J8'); ws['A8'].value = '  귀하께서 의뢰하신 건에 대하여 하기와 같이 견적합니다.'
    ws['A8'].font = fn(10,False,'444444'); ws['A8'].fill = fl('F6F9F6')

    # Row 9: 총금액 (수식으로 연결)
    ws.merge_cells('A9:D9'); ws['A9'].value = '총 금 액 (VAT 포함)'; ws['A9'].font = Font(name='맑은 고딕',size=13,bold=True,color='F0D080'); ws['A9'].fill = fl(DG); ws['A9'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[9].height = 30

    # Row 10: 테이블 헤더
    hdrs = [('A','No.'),('B','품목명 / 시공내용'),('C','규격'),('D','단위'),('E','수량'),('F','단가(원)'),('G','공급가액'),('H','부가세(10%)'),('I','합계'),('J','비고')]
    for col, hdr in hdrs:
        ws[f'{col}10'].value = hdr; ws[f'{col}10'].font = fn(10,True,'FFFFFF')
        ws[f'{col}10'].fill = fl(GOLD if col == 'I' else DG)
        ws[f'{col}10'].alignment = Alignment(horizontal='center', vertical='center'); ws[f'{col}10'].border = bd()
    ws.row_dimensions[10].height = 24

    # Row 11~: 항목 (실제 견적 데이터 미리 채움)
    items = []
    items.append((spec_info_text, f'가로 {width_input}m', 'm', 1, base_price, '어닝 본체'))
    if fabric_add_price > 0: items.append((f'원단 추가 ({fabric_type})', '', '식', 1, fabric_add_price, ''))
    if use_print and print_price > 0: items.append(('로고 인쇄 (1도/1면)', '', '식', 1, print_price, ''))
    if use_light and light_price > 0: items.append(('T5 조명', '', '개', light_count, 20000, ''))
    if use_gutter and gutter_price > 0: items.append(('물받이 설치', '', '식', 1, gutter_price, ''))
    if use_remove and remove_price > 0: items.append(('기존 철거', '', '식', 1, remove_price, ''))
    if use_ladder and ladder_price > 0: items.append(('장비 사용 (스카이)', '', '식', 1, ladder_price, ''))
    if use_bracket and bracket_price > 0: items.append(('특수 브라켓', '', '개', bracket_count, 20000, ''))
    if use_special_wall and special_wall_price > 0: items.append(('특수 벽면 시공', '', '식', 1, special_wall_price, ''))
    if use_pole and pole_price > 0: items.append(('보조 기둥 설치', '', '식', 1, pole_price, ''))
    if use_night and night_price > 0: items.append(('야간/주말 할증', '', '식', 1, night_price, ''))
    items.append(('시공 노무비', '', '식', 1, labor_price, ''))
    if material_price > 0: items.append(('부자재비용', '', '식', 1, material_price, ''))
    if transport_price > 0: items.append(('운송비용', '', '식', 1, transport_price, ''))

    # 여유 빈 행 3개
    for _ in range(3):
        items.append(('', '', '', None, None, ''))

    fr = 11  # 첫 데이터행
    for i, (name, spec, unit, qty_val, price_val, note) in enumerate(items):
        r = 11 + i
        bg = fl('FFFFFF') if i % 2 == 0 else fl(LGRY)
        is_empty = (name == '')

        ws[f'A{r}'].value = i+1 if not is_empty else ''; ws[f'A{r}'].font = fn(9); ws[f'A{r}'].fill = bg; ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center'); ws[f'A{r}'].border = bd()
        ws[f'B{r}'].value = name; ws[f'B{r}'].font = fn(9); ws[f'B{r}'].fill = bg; ws[f'B{r}'].border = bd()
        ws[f'C{r}'].value = spec; ws[f'C{r}'].font = fn(8,False,'555555'); ws[f'C{r}'].fill = bg; ws[f'C{r}'].alignment = Alignment(horizontal='center', vertical='center'); ws[f'C{r}'].border = bd()
        ws[f'D{r}'].value = unit; ws[f'D{r}'].font = fn(9); ws[f'D{r}'].fill = bg; ws[f'D{r}'].alignment = Alignment(horizontal='center', vertical='center'); ws[f'D{r}'].border = bd()

        # E(수량), F(단가) - 노란 입력 셀
        ws[f'E{r}'].value = qty_val; ws[f'E{r}'].fill = fl(CRM) if not is_empty else bg; ws[f'E{r}'].font = fn(10,True); ws[f'E{r}'].alignment = Alignment(horizontal='center', vertical='center'); ws[f'E{r}'].border = bd(); ws[f'E{r}'].number_format = '#,##0'
        ws[f'F{r}'].value = price_val; ws[f'F{r}'].fill = fl(CRM) if not is_empty else bg; ws[f'F{r}'].font = fn(10,True); ws[f'F{r}'].alignment = Alignment(horizontal='right', vertical='center'); ws[f'F{r}'].border = bd(); ws[f'F{r}'].number_format = '#,##0'

        # G(공급가), H(부가세), I(합계) - 자동 수식
        ws[f'G{r}'].value = f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r})),E{r}*F{r},"")'; ws[f'G{r}'].font = fn(10,True); ws[f'G{r}'].fill = bg; ws[f'G{r}'].alignment = Alignment(horizontal='right', vertical='center'); ws[f'G{r}'].border = bd(); ws[f'G{r}'].number_format = '#,##0'
        ws[f'H{r}'].value = f'=IF(ISNUMBER(G{r}),INT(G{r}*0.1),"")'; ws[f'H{r}'].font = fn(10); ws[f'H{r}'].fill = bg; ws[f'H{r}'].alignment = Alignment(horizontal='right', vertical='center'); ws[f'H{r}'].border = bd(); ws[f'H{r}'].number_format = '#,##0'
        ws[f'I{r}'].value = f'=IF(ISNUMBER(G{r}),G{r}+H{r},"")'; ws[f'I{r}'].font = Font(name='맑은 고딕',size=10,bold=True,color=DG); ws[f'I{r}'].fill = fl(LGOL) if not is_empty else bg; ws[f'I{r}'].alignment = Alignment(horizontal='right', vertical='center'); ws[f'I{r}'].border = bd(); ws[f'I{r}'].number_format = '#,##0'
        ws[f'J{r}'].value = note; ws[f'J{r}'].font = fn(8,False,'666666'); ws[f'J{r}'].fill = bg; ws[f'J{r}'].border = bd()
        ws.row_dimensions[r].height = 22

    lr = 11 + len(items) - 1  # 마지막 데이터행

    # 합계 섹션
    cur = lr + 2
    for lbl, is_total in [('공급가액 합계 (VAT 제외)', False), ('부가세 합계 (10%)', False), ('총 합 계 금 액 (VAT 포함)', True)]:
        ws.merge_cells(f'A{cur}:H{cur}'); ws[f'A{cur}'].value = lbl
        ws[f'A{cur}'].font = Font(name='맑은 고딕', size=13 if is_total else 10, bold=True, color='F0D080' if is_total else 'FFFFFF')
        ws[f'A{cur}'].fill = fl(DG if is_total else GOLD); ws[f'A{cur}'].alignment = Alignment(horizontal='right', vertical='center'); ws[f'A{cur}'].border = bd()
        ws.merge_cells(f'I{cur}:J{cur}')
        c = ws[f'I{cur}']
        if is_total:
            c.value = f'=I{cur-2}+I{cur-1}'
            total_row = cur
        elif 'VAT' in lbl:
            c.value = ArrayFormula(f'I{cur}', 'S' + f'SUMPRODUCT(IFERROR(G{fr}:G{lr}*1,0))')
        else:
            c.value = ArrayFormula(f'I{cur}', 'S' + f'SUMPRODUCT(IFERROR(H{fr}:H{lr}*1,0))')
        c.font = Font(name='맑은 고딕', size=16 if is_total else 12, bold=True, color=DG)
        c.fill = fl(LGOL); c.alignment = Alignment(horizontal='right', vertical='center'); c.border = bd(); c.number_format = '#,##0'
        ws.row_dimensions[cur].height = 30 if is_total else 24
        cur += 1

    # E9 총금액 연결
    ws.merge_cells('E9:J9')
    ws['E9'].value = f'=I{total_row}'; ws['E9'].font = Font(name='맑은 고딕',size=16,bold=True,color=DG)
    ws['E9'].fill = fl(LGOL); ws['E9'].alignment = Alignment(horizontal='right', vertical='center'); ws['E9'].number_format = '#,##0"원 정"'

    # 한글 금액
    ws.merge_cells(f'A{cur}:J{cur}')
    ws[f'A{cur}'].value = f'=IF(I{total_row}="","","일금  "&NUMBERSTRING(I{total_row},1)&"원정  (부가세 포함)")'
    ws[f'A{cur}'].font = Font(name='맑은 고딕',size=12,bold=True,color=DG); ws[f'A{cur}'].fill = fl(SAGE); ws[f'A{cur}'].alignment = Alignment(horizontal='center', vertical='center')
    cur += 2

    # 특기사항
    ws.merge_cells(f'A{cur}:J{cur}'); ws[f'A{cur}'].value = '■  특 기 사 항'; ws[f'A{cur}'].font = fn(11,True,'FFFFFF'); ws[f'A{cur}'].fill = fl(DG)
    cur += 1
    notes = [
        '① 결제: 계약금 50%(계약시, 입금 후 시공착수) / 중도금 30%(자재반입후 3일이내) / 잔금 20%(공사완료후 7일이내).',
        '② 하자보증: 시공완료일로부터 1년 무상 A/S (천재지변·사용자 과실·임의 개조·소모성 부품·자연 열화 제외).',
        '③ 지체상금: 공사 지연 시 지체일수 × 계약금액의 1/100 적용 (천재지변·불가항력 제외).',
        '④ 면책: 본 견적은 개략 견적이며 실측 후 변동 가능 / 장비·지반·인허가 비용 별도 / 건축주 귀책 중단 시 기투입 비용 미반환.',
        '⑤ 원자재 가격·환율 변동 시 단가 조정 가능 / 불가항력 사유 시 납기 변경 가능.',
        '⑥ 세금계산서: 부가세 포함 금액 기준 발행 가능 (요청 시).',
        '⑦ 문의: 010-4337-0582 | candlejs6@gmail.com | wocs.kr',
    ]
    for j, note in enumerate(notes):
        ws.merge_cells(f'A{cur}:J{cur}'); ws[f'A{cur}'].value = f' {note}'
        ws[f'A{cur}'].font = fn(9); ws[f'A{cur}'].fill = fl(SAGE if j%2==0 else 'FFFFFF')
        ws.row_dimensions[cur].height = 18; cur += 1

    # 서명란
    cur += 1
    ws.merge_cells(f'A{cur}:J{cur}'); ws[f'A{cur}'].value = '위 견적 내용을 확인하였으며 이에 동의합니다.'
    ws[f'A{cur}'].font = fn(10); ws[f'A{cur}'].alignment = Alignment(horizontal='center', vertical='center')
    cur += 1
    ws.merge_cells(f'A{cur}:B{cur}'); ws[f'A{cur}'].value = '공급자'; ws[f'A{cur}'].font = fn(10,True,'FFFFFF'); ws[f'A{cur}'].fill = fl(DG); ws[f'A{cur}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'C{cur}:F{cur}'); ws[f'C{cur}'].value = '우성어닝천막공사캠프시스템   김우성'; ws[f'C{cur}'].font = fn(10)
    ws.merge_cells(f'G{cur}:H{cur}'); ws[f'G{cur}'].value = '수급자'; ws[f'G{cur}'].font = fn(10,True,'FFFFFF'); ws[f'G{cur}'].fill = fl(DG); ws[f'G{cur}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'I{cur}:J{cur}'); ws[f'I{cur}'].value = '                              (인)'; ws[f'I{cur}'].font = fn(10)

    # 도장 이미지 삽입
    stamp_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'stamp_clean.png')
    if os.path.exists(stamp_path):
        from openpyxl.drawing.image import Image as XlImage
        stamp_img = XlImage(stamp_path)
        stamp_img.width = 45
        stamp_img.height = 50
        ws.add_image(stamp_img, f'F{cur}')

    ws.print_area = f'A1:J{cur}'
    ws.page_setup.orientation = 'portrait'; ws.page_setup.paperSize = 9; ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0

    if wb.calculation is not None:
        wb.calculation.calcId = 0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

st.write("")
st.write("")
col_dn1, col_dn2, col_dn3, col_dn4 = st.columns([2, 1, 1, 1])
with col_dn2:
    st.download_button("💾 이미지 저장", create_image(), f"고정식견적_{customer_name}_{today_str}.png", "image/png", use_container_width=True)
with col_dn3:
    st.download_button("📊 Excel 저장", create_excel(), f"고정식견적_{customer_name}_{today_str}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
with col_dn4:
    st.download_button(
        "📋 정식 견적서",
        create_formal_excel(),
        f"WOCS_정식견적_고정어닝_{customer_name}_{today_str}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="v7 Excel 수식 포함 정식 견적서 (수량·단가 수정 → 자동 계산)"
    )