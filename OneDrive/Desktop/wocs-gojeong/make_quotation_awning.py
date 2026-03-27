#!/usr/bin/env python3
"""
WOCS 천막·어닝 구조물 전용 견적서 생성기
────────────────────────────────────────────────────────────
사용법:
  python3 make_quotation.py                    # 대화형 입력
  python3 make_quotation.py --type awning      # 천막·어닝 견적서
  python3 make_quotation.py --type glamping    # 글램핑 텐트 견적서
  python3 make_quotation.py --list             # 저장된 견적 목록 출력
────────────────────────────────────────────────────────────
"""

import os, sys, json, argparse
from datetime import date, timedelta
from pathlib import Path

# ── 의존성 체크 ──
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl --break-system-packages")
    sys.exit(1)

# PDF 변환 (선택 — 없어도 xlsx 생성 가능)
try:
    import subprocess
    HAS_LIBREOFFICE = subprocess.run(
        ['which', 'libreoffice'], capture_output=True).returncode == 0
except:
    HAS_LIBREOFFICE = False

# ── 경로 설정 ──
BASE_DIR    = Path(__file__).parent
OUTPUT_DIR  = BASE_DIR / 'output'
COUNTER_FILE = BASE_DIR / '.quote_counter.json'
OUTPUT_DIR.mkdir(exist_ok=True)

# ══════════════════════════════════════════════════════════
# 색상 팔레트 (고급 포레스트 그린 + 샴페인 골드)
# ══════════════════════════════════════════════════════════
C = {
    'DG'  : '1A3526',  # 딥 포레스트 그린 (헤더)
    'GOLD': 'B59035',  # 샴페인 골드 (합계)
    'SAGE': 'E8F0EA',  # 세이지 그린 (섹션)
    'CRM' : 'FDF8EE',  # 아이보리 크림 (입력)
    'LGRY': 'F4F6F4',  # 그린틴트 (짝수행)
    'W'   : 'FFFFFF',
    'DK'  : '1A1A1A',
    'GTXT': 'F0D080',  # 소프트 골드 텍스트
    'MDGN': '2D5A3D',  # 섹션 제목 텍스트
    'LGOL': 'FDF3DC',  # 합계값 배경
}

# ══════════════════════════════════════════════════════════
# 스타일 헬퍼
# ══════════════════════════════════════════════════════════
def fl(c):
    return PatternFill('solid', start_color=c, end_color=c)

def fn(sz=10, bold=False, color=None):
    return Font(name='맑은 고딕', size=sz, bold=bold, color=color or C['DK'])

def bd():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def ca(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def la(i=1): return Alignment(horizontal='left',  vertical='center', wrap_text=True, indent=i)
def ra(): return Alignment(horizontal='right', vertical='center', wrap_text=False)

def sc(ws, addr, val=None, font=None, fill=None, align=None, border=None, fmt=None):
    """set_cell — 셀 속성 한 번에 설정"""
    c = ws[addr]
    if val   is not None: c.value  = val
    if font  is not None: c.font   = font
    if fill  is not None: c.fill   = fill
    if align is not None: c.alignment = align
    if border is not None: c.border = border
    if fmt   is not None: c.number_format = fmt

def mc(ws, rng):
    """병합 — 이미 병합된 경우 해제 후 재병합"""
    try: ws.unmerge_cells(rng)
    except: pass
    ws.merge_cells(rng)

def apply_merge_border(ws, rng):
    """병합 범위 전체에 방향별 테두리 적용"""
    from openpyxl.utils import range_boundaries
    try: c1, r1, c2, r2 = range_boundaries(rng)
    except: return
    s = Side(style='thin'); n = Side(style=None)
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = Border(
                left   = s if c==c1 else n,
                right  = s if c==c2 else n,
                top    = s if r==r1 else n,
                bottom = s if r==r2 else n,
            )

# ══════════════════════════════════════════════════════════
# 견적번호 자동 채번
# ══════════════════════════════════════════════════════════
def get_next_quote_number(quote_type: str) -> str:
    """WOCS-YYYY-{TYPE}-NNN 형식으로 자동 채번"""
    year = date.today().year
    prefix = 'A' if quote_type == 'awning' else 'G'
    key = f"{year}_{prefix}"

    data = {}
    if COUNTER_FILE.exists():
        try:
            data = json.loads(COUNTER_FILE.read_text(encoding='utf-8'))
        except:
            data = {}

    seq = data.get(key, 0) + 1
    data[key] = seq
    COUNTER_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                            encoding='utf-8')
    return f"WOCS-{year}-{prefix}{seq:03d}"

def list_quotes():
    """저장된 견적 목록 출력"""
    files = sorted(OUTPUT_DIR.glob('*.xlsx'))
    if not files:
        print("저장된 견적서가 없습니다.")
        return
    print(f"\n{'─'*60}")
    print(f"{'파일명':<40} {'크기':>8}  {'수정일'}")
    print(f"{'─'*60}")
    for f in files:
        stat = f.stat()
        mtime = date.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d')
        size  = f"{stat.st_size/1024:.1f}KB"
        print(f"  {f.name:<38} {size:>8}  {mtime}")
    print(f"{'─'*60}")
    print(f"총 {len(files)}개\n")

# ══════════════════════════════════════════════════════════
# 품목 데이터
# ══════════════════════════════════════════════════════════

AWNING_SECTIONS = [
    ('자재비 ①  원단·방수포류', [
        ('PVC 방수포 — 타포린 진회색 (900g/m²)', '0.9mm 두께',   '야드', '내구성 최상급'),
        ('PVC 방수포 — 타포린 흰색 (900g/m²)',   '0.9mm 두께',   '야드', '투명도 높음'),
        ('방염 타포린 (KFI 방염 인증)',           'KFI 인증',     '야드', '관공서·학교 필수'),
        ('아스테이지 — 투명 PVC 측면막',          '0.3mm',        '야드', '측면 방풍'),
        ('방수포 재단·고주파 제작비',             '봉제+아일렛',  '야드', ''),
        ('오비띠 — 천막 고정 보강 벨트',         '50mm 폭',      'm',   '내부 보강'),
        ('로프 및 아일렛',                       '4mm PP',       'm',   '고정용'),
    ]),
    ('자재비 ②  철골·파이프 구조물', [
        ('각파이프 (아연도금)',                   '40×40×1.8T',  '본',   '기본 프레임'),
        ('각파이프 (아연도금)',                   '50×50×2.3T',  '본',   '대형 구조용'),
        ('원형파이프 기둥 (아연도금)',             'Ø60.5×2.3T',  '본',   '기둥재'),
        ('H빔 — 철골 주 구조재',                 '100×100×6T',  'm',    '대형 차양'),
        ('플랫바 — 연결 철물',                   '6T×40×L',     '개',   '조인트 연결'),
        ('앵커 볼트 세트',                       'M12×200',     '세트', '기초 고정'),
        ('설치부자재 일체 (볼트·너트·와셔)',       '',            '식',   ''),
    ]),
    ('자재비 ③  부속·마감재 (선택)', [
        ('알루미늄 커튼레일 — 이동식',            '알루미늄',     'm',   '슬라이딩 개폐'),
        ('레일바퀴 (나일론)',                     '40mm',        '개',   '레일용'),
        ('폴리카보네이트 패널 — 지붕재',          '5mm 투명',    'm²',  '채광 가능'),
        ('방충망 (하절기 선택)',                  '18×16 mesh',  'm²',  ''),
        ('실리콘 코킹 — 방수 마감',               '비초산형',    '식',   '방수 마감'),
        ('방청 도장 — 철골 녹 방지',              '2회 도장',    '식',   '구조물 보호'),
    ]),
    ('시공비', [
        ('현장 실측 및 설계',                    '',            '식',   '견적 전 필수'),
        ('기초 앵커·콘크리트 공사',              '토질별 조건',  '개소', '지반 조건별'),
        ('철골 구조물 용접·조립',               '',            '식',   ''),
        ('설치 인건비 (2인 1팀)',               '',            '일',   '출장비 포함'),
        ('방청·마감 도장 인건비',              '',            '식',   '구조물 도장'),
        ('폐자재 정리·처리비',                 '',            '식',   '현장 클린'),
        ('운반비 (전국 직시공)',               '',            '식',   ''),
        ('공과잡비',                           '',            '식',   '소모품 일체'),
    ]),
]


AWNING_NOTES = [
    ' ① 노란 셀 (E:수량, F:단가) 에 숫자 입력 → 공급가·부가세·합계·총금액 자동 계산.',
    ' ② 결제: 계약금 50%(계약시, 입금 완료 후 시공 착수) / 중도금 30%(자재 반입 후 3일 이내) / 잔금 20%(공사 완료 후 7일 이내).',
    ' ③ 하자보증: 시공 완료일로부터 1년 무상 A/S (천재지변·사용자 과실·임의 개조·소모성 부품·자연 열화 제외).',
    ' ④ 지체상금: 공사 지연 시 지체일수 × 계약금액의 1/100 적용 (천재지변·불가항력 제외).',
    ' ⑤ 면책: 본 견적은 개략 견적이며 실측 후 변동 가능 / 지반·장비·인허가 비용 별도 / 건축주 귀책 중단 시 기투입 비용 미반환.',
    ' ⑥ 원자재 가격·환율 변동 시 단가 조정 가능 / 불가항력 사유 시 납기 변경 가능.',
    ' ⑦ 세금계산서: 부가세 포함 금액 기준 발행 가능 (요청 시).',
    ' ⑧ 문의: 010-4337-0582 | candlejs6@gmail.com | wocs.kr',
]


# ══════════════════════════════════════════════════════════
# 시트 빌더
# ══════════════════════════════════════════════════════════
def build_sheet(ws, info: dict, sections: list, notes: list):
    """
    info = {
      'quote_no': str,
      'company': str,
      'biz_type': str,
      'item_desc': str,
      'customer': str,
      'address': str,
      'phone': str,
      'work_date': str,
    }
    """
    ws.sheet_view.showGridLines = False
    for col, w in {'A':5,'B':26,'C':12,'D':5,'E':8,'F':13,
                   'G':14,'H':12,'I':14,'J':13}.items():
        ws.column_dimensions[col].width = w

    # ── Row 1: 타이틀 ──
    ws.row_dimensions[1].height = 44
    mc(ws, 'A1:J1')
    sc(ws,'A1','見   積   書',
       Font(name='맑은 고딕',size=26,bold=True,color=C['GTXT']),
       fl(C['DG']), ca(), bd())

    # ── Row 2: 견적번호/날짜 ──
    ws.row_dimensions[2].height = 22
    for rng in ['A2:B2','C2:D2','E2:F2','G2:I2','J2:J2']:
        mc(ws, rng); apply_merge_border(ws, rng)
    sc(ws,'A2','견  적  번  호', fn(9,True,C['W']), fl(C['DG']), ca(), bd())
    sc(ws,'C2', info['quote_no'], fn(10,False,C['DK']), fl(C['CRM']), la(), bd())
    sc(ws,'E2','작  성  일',     fn(9,True,C['W']), fl(C['DG']), ca(), bd())
    sc(ws,'G2','=TEXT(TODAY(),"YYYY년 MM월 DD일")',
       fn(10,True,C['DG']), fl(C['CRM']), ra(), bd())
    sc(ws,'J2','유효: 10일',     fn(9,False,'888888'), fl(C['W']), ca(), bd())

    # ── Rows 3~7: 고객정보 + 공급자 ──
    for r in range(3,8): ws.row_dimensions[r].height = 22
    mc(ws,'F3:F7'); apply_merge_border(ws,'F3:F7')
    sc(ws,'F3','공  급  자', fn(10,True,C['W']), fl(C['DG']), ca(), bd())

    sup_rows = [
        ('사업자\n등록번호', '465-02-03270',      '',       ''),
        ('상  호',          info['company'],      '대표자', '김우성'),
        ('사업장소재지',    '전남 화순군 사평면 유마로 592','',''),
        ('업  태',         info['biz_type'],      '종  목', info['item_desc']),
        ('전  화',         '010-4337-0582',       '사이트', 'wocs.kr'),
    ]
    for i,(l1,v1,l2,v2) in enumerate(sup_rows):
        r = 3+i
        sc(ws,f'G{r}',l1, fn(9,True,C['W']), fl(C['DG']), ca(), bd())
        mc(ws,f'H{r}:I{r}'); apply_merge_border(ws,f'H{r}:I{r}')
        sc(ws,f'H{r}',v1, fn(8 if r==5 else 9), fl(C['W']), la(), bd())
        sc(ws,f'J{r}',f'{l2}: {v2}' if l2 else '',
           fn(9,False,'333333'), fl(C['W']), la(1), bd())

    mc(ws,'A3:E3'); apply_merge_border(ws,'A3:E3')
    sc(ws,'A3','■  수  신  /  고  객  정  보',
       fn(10,True,C['W']), fl(C['DG']), la(), bd())

    cust_data = [
        (4,'고 객 명 (업체명)', info.get('customer','')),
        (5,'현 장 주 소',       info.get('address', '')),
        (6,'연 락 처',          info.get('phone',   '')),
        (7,'시공 희망일',       info.get('work_date','')),
    ]
    for r, lbl, val in cust_data:
        sc(ws,f'A{r}',lbl, fn(9,True,C['W']), fl(C['DG']), ca(), bd())
        mc(ws,f'B{r}:E{r}'); apply_merge_border(ws,f'B{r}:E{r}')
        sc(ws,f'B{r}',val, fn(9), fl(C['CRM']), la(), bd())

    # ── Row 8: 견적 문구 ──
    ws.row_dimensions[8].height = 18
    mc(ws,'A8:J8'); apply_merge_border(ws,'A8:J8')
    sc(ws,'A8','   귀하께서 의뢰하신 건에 대하여 하기와 같이 견적합니다.',
       fn(10,False,'444444'), fl('F6F9F6'), la())

    # ── Row 9: 납품/결재/보증 ──
    ws.row_dimensions[9].height = 22
    for rng in ['A9:B9','C9:D9','E9:F9','G9:H9','I9:J9']:
        mc(ws,rng); apply_merge_border(ws,rng)
    sc(ws,'A9','납품 / 시공일', fn(9,True,C['W']), fl(C['DG']), ca(), bd())
    sc(ws,'C9',info.get('work_date',''), fn(9), fl(C['CRM']), ca(), bd())
    sc(ws,'E9','결  재  방  법', fn(9,True,C['W']), fl(C['DG']), ca(), bd())
    sc(ws,'G9','계약금 50% → 잔금 50%', fn(9), fl(C['W']), ca(), bd())
    sc(ws,'I9','보증기간: 납품 후 1년',  fn(9), fl(C['W']), ca(), bd())

    # ── Row 10: 총금액 ──
    ws.row_dimensions[10].height = 30
    mc(ws,'A10:D10'); mc(ws,'E10:J10')
    apply_merge_border(ws,'A10:D10'); apply_merge_border(ws,'E10:J10')
    sc(ws,'A10','총    금    액    (VAT 포함)',
       Font(name='맑은 고딕',size=13,bold=True,color=C['GTXT']),
       fl(C['DG']), ca(), bd())
    # 총금액 수식은 합계행 확정 후 삽입

    # ── Row 11: 골드 구분선 ──
    ws.row_dimensions[11].height = 4
    mc(ws,'A11:J11'); ws['A11'].fill = fl(C['GOLD'])

    # ── Row 12: 단위 표시 ──
    ws.row_dimensions[12].height = 10
    mc(ws,'H12:J12')
    sc(ws,'H12','(단위 : 원 / 야드 / m / m² / 본 / 개 / 식)',
       fn(8,False,'888888'), align=ra())

    # ── Row 13: 테이블 헤더 ──
    ws.row_dimensions[13].height = 26
    hdrs = [
        ('A','No.',False),('B','품  목  명  /  시  공  내  용',False),
        ('C','규    격',False),('D','단위',False),('E','수  량',False),
        ('F','단  가 (원)',False),('G','공  급  가  액',False),
        ('H','부가세 (10%)',False),('I','합    계',True),('J','비  고',False),
    ]
    for col,hdr,gold in hdrs:
        sc(ws,f'{col}13',hdr,
           fn(10,True,C['W']),
           fl(C['GOLD'] if gold else C['DG']), ca(), bd())

    # ── 품목 행 ──
    cur = 14
    all_rows = []
    item_no = 1

    for sec_title, items in sections:
        ws.row_dimensions[cur].height = 21
        mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
        sc(ws,f'A{cur}',f'  ▶  {sec_title}',
           fn(10,True,C['MDGN']), fl(C['SAGE']), la(2), bd())
        cur += 1

        for name, spec, unit, note in items:
            r = cur
            ws.row_dimensions[r].height = 22
            is_empty = (name == '')
            bg = fl(C['W']) if len(all_rows) % 2 == 0 else fl(C['LGRY'])

            sc(ws,f'A{r}', item_no if not is_empty else '', fn(9), bg, ca(), bd())
            sc(ws,f'B{r}', name, fn(9), bg, la(1), bd())
            sc(ws,f'C{r}', spec, fn(8,False,'555555'), bg, ca(), bd())
            sc(ws,f'D{r}', unit, fn(9), bg, ca(), bd())

            for col_idx, col_letter in [(5,'E'),(6,'F')]:
                c = ws[f'{col_letter}{r}']
                c.fill = fl(C['CRM']) if not is_empty else bg
                c.font = fn(10,True,C['DK'])
                c.alignment = ca() if col_letter == 'E' else ra()
                c.border = bd()
                c.number_format = '#,##0'

            c = ws[f'G{r}']
            c.value = f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r})),E{r}*F{r},"")' if not is_empty else ''
            c.font = fn(10,True); c.fill = bg; c.alignment = ra()
            c.border = bd(); c.number_format = '#,##0'

            c = ws[f'H{r}']
            c.value = f'=IF(ISNUMBER(G{r}),INT(G{r}*0.1),"")' if not is_empty else ''
            c.font = fn(10); c.fill = bg; c.alignment = ra()
            c.border = bd(); c.number_format = '#,##0'

            c = ws[f'I{r}']
            c.value = f'=IF(ISNUMBER(G{r}),G{r}+H{r},"")' if not is_empty else ''
            c.font = Font(name='맑은 고딕',size=10,bold=True,
                          color=C['DG'] if not is_empty else C['DK'])
            c.fill = fl(C['LGOL']) if not is_empty else bg
            c.alignment = ra(); c.border = bd(); c.number_format = '#,##0'

            sc(ws,f'J{r}', note, fn(8,False,'666666'), bg, la(1), bd())

            all_rows.append(r)
            if not is_empty: item_no += 1
            cur += 1

    # 여유 빈 행 3개
    for _ in range(3):
        r = cur; ws.row_dimensions[r].height = 22
        bg = fl(C['W']) if cur % 2 == 0 else fl(C['LGRY'])
        for col in range(1, 11):
            cell = ws.cell(row=r, column=col)
            cell.fill = fl(C['CRM']) if col in (5,6) else bg
            cell.border = bd()
            cell.number_format = '#,##0'
        ws.cell(r,7).value = f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r})),E{r}*F{r},"")'
        ws.cell(r,7).font = fn(10,True); ws.cell(r,7).fill = bg
        ws.cell(r,7).alignment = ra(); ws.cell(r,7).number_format = '#,##0'
        ws.cell(r,8).value = f'=IF(ISNUMBER(G{r}),INT(G{r}*0.1),"")'
        ws.cell(r,8).font = fn(10); ws.cell(r,8).fill = bg
        ws.cell(r,8).alignment = ra(); ws.cell(r,8).number_format = '#,##0'
        ws.cell(r,9).value = f'=IF(ISNUMBER(G{r}),G{r}+H{r},"")'
        ws.cell(r,9).font = Font(name='맑은 고딕',size=10,bold=True,color=C['DG'])
        ws.cell(r,9).fill = fl(C['LGOL']); ws.cell(r,9).alignment = ra()
        ws.cell(r,9).border = bd(); ws.cell(r,9).number_format = '#,##0'
        all_rows.append(r); cur += 1

    fr = all_rows[0]; lr = all_rows[-1]

    # ── 합계 섹션 ──
    ws.row_dimensions[cur].height = 5; cur += 1

    for lbl, formula, is_total in [
        ('공  급  가  액  합  계  (VAT 제외)',
         f'=SUMPRODUCT(IFERROR(G{fr}:G{lr}*1,0))', False),
        ('부  가  세  합  계  (10%)',
         f'=SUMPRODUCT(IFERROR(H{fr}:H{lr}*1,0))', False),
        ('총  합  계  금  액  (VAT 포함)',
         None, True),
    ]:
        ws.row_dimensions[cur].height = 30 if is_total else 24
        mc(ws,f'A{cur}:H{cur}'); apply_merge_border(ws,f'A{cur}:H{cur}')
        sc(ws,f'A{cur}',lbl,
           Font(name='맑은 고딕',size=13 if is_total else 10,bold=True,
                color=C['GTXT'] if is_total else C['W']),
           fl(C['DG'] if is_total else C['GOLD']),
           Alignment(horizontal='right',vertical='center'), bd())
        mc(ws,f'I{cur}:J{cur}'); apply_merge_border(ws,f'I{cur}:J{cur}')
        c = ws[f'I{cur}']
        if is_total:
            c.value = f'=I{cur-2}+I{cur-1}'
        else:
            # v7과 동일하게 배열 수식(Array Formula)으로 저장
            from openpyxl.worksheet.formula import ArrayFormula
            c.value = ArrayFormula(f'I{cur}', formula[1] + formula[1:])  # openpyxl 3.x 버그 우회: 첫글자 중복
        c.font = Font(name='맑은 고딕',size=16 if is_total else 12,bold=True,
                      color=C['DG'] if is_total else C['DK'])
        c.fill = fl(C['LGOL']); c.alignment = ra()
        c.border = bd(); c.number_format = '#,##0'
        if is_total:
            total_row = cur
            ws['E10'].value = f'=I{total_row}'
            ws['E10'].font  = Font(name='맑은 고딕',size=16,bold=True,color=C['DG'])
            ws['E10'].fill  = fl(C['LGOL']); ws['E10'].alignment = ra()
            ws['E10'].border = bd(); ws['E10'].number_format = '#,##0"원 정"'
        cur += 1

    # 한글 금액
    ws.row_dimensions[cur].height = 22
    mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
    c = ws[f'A{cur}']
    c.value = (f'=IF(I{total_row}="","","일금  "&NUMBERSTRING(I{total_row},1)'
               f'&"원정  (부가세 포함)")')
    c.font = Font(name='맑은 고딕',size=12,bold=True,color=C['DG'])
    c.fill = fl(C['SAGE']); c.alignment = ca(); c.border = bd()
    cur += 1

    # ── 특기사항 ──
    ws.row_dimensions[cur].height = 5; cur += 1
    ws.row_dimensions[cur].height = 24
    mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
    sc(ws,f'A{cur}','■  특  기  사  항',
       fn(11,True,C['W']), fl(C['DG']), la(), bd()); cur += 1

    for j, note in enumerate(notes):
        ws.row_dimensions[cur].height = 18
        mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
        sc(ws,f'A{cur}',note, fn(9),
           fl(C['SAGE'] if j%2==0 else C['W']), la(), bd())
        cur += 1

    # ── 서명란 ──
    ws.row_dimensions[cur].height = 5; cur += 1
    ws.row_dimensions[cur].height = 32
    mc(ws,f'A{cur}:J{cur}'); apply_merge_border(ws,f'A{cur}:J{cur}')
    sc(ws,f'A{cur}','위 견적 내용을 확인하였으며 이에 동의합니다.',
       fn(10), fl(C['W']), ca(), bd()); cur += 1

    ws.row_dimensions[cur].height = 46
    mc(ws,f'A{cur}:B{cur}'); apply_merge_border(ws,f'A{cur}:B{cur}')
    sc(ws,f'A{cur}','공  급  자', fn(10,True,C['W']), fl(C['DG']), ca(), bd())
    mc(ws,f'C{cur}:F{cur}'); apply_merge_border(ws,f'C{cur}:F{cur}')
    sc(ws,f'C{cur}',f"{info['company']}   김우성", fn(10), fl(C['W']), ca(), bd())
    # 도장 이미지 삽입
    import os as _os
    _stamp_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'stamp_clean.png')
    if _os.path.exists(_stamp_path):
        from openpyxl.drawing.image import Image as XlImage
        _stamp = XlImage(_stamp_path)
        _stamp.width = 45
        _stamp.height = 50
        ws.add_image(_stamp, f'F{cur}')
    mc(ws,f'G{cur}:H{cur}'); apply_merge_border(ws,f'G{cur}:H{cur}')
    sc(ws,f'G{cur}','수  급  자', fn(10,True,C['W']), fl(C['DG']), ca(), bd())
    mc(ws,f'I{cur}:J{cur}'); apply_merge_border(ws,f'I{cur}:J{cur}')
    sc(ws,f'I{cur}','                              (인)', fn(10), fl(C['W']), ca(), bd())

    # 인쇄 설정
    ws.print_area = f'A1:J{cur}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0


# ══════════════════════════════════════════════════════════
# 천막·어닝 견적서 생성
# ══════════════════════════════════════════════════════════
def generate_awning_quote(info: dict) -> Path:
    wb = Workbook()
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = '천막·어닝 구조물 견적서'
    info1 = dict(info)
    info1['company']   = '우성어닝천막공사캠프시스템'
    info1['biz_type']  = '제조업, 건설업'
    info1['item_desc'] = '천막·어닝·방수포·구조물'
    build_sheet(ws, info1, AWNING_SECTIONS, AWNING_NOTES)
    today_str = date.today().strftime('%Y%m%d')
    cust_safe = info.get('customer', '').replace(' ','_').replace('/','_') or '견적'
    filename  = f"{today_str}_{info['quote_no']}_{cust_safe}.xlsx"
    out_path  = OUTPUT_DIR / filename
    if wb.calculation is not None:
        wb.calculation.calcId = 0
    wb.save(str(out_path))
    return out_path


def generate_awning_bytes(info: dict) -> bytes:
    import io as _io
    wb = Workbook()
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = '천막·어닝 구조물 견적서'
    info1 = dict(info)
    info1['company']   = '우성어닝천막공사캠프시스템'
    info1['biz_type']  = '제조업, 건설업'
    info1['item_desc'] = '천막·어닝·방수포·구조물'
    build_sheet(ws, info1, AWNING_SECTIONS, AWNING_NOTES)
    buf = _io.BytesIO()
    if wb.calculation is not None:
        wb.calculation.calcId = 0
    wb.save(buf)
    return buf.getvalue()


if __name__ == '__main__':
    quote_no = get_next_quote_number('awning')
    print(f"견적번호: {quote_no}")
    customer = input("고객명: ").strip()
    address  = input("현장주소: ").strip()
    phone    = input("연락처: ").strip()
    work_date = input("시공희망일 (Enter=미정): ").strip() or '미정'
    info = {'quote_no': quote_no, 'customer': customer,
            'address': address, 'phone': phone, 'work_date': work_date}
    path = generate_awning_quote(info)
    print(f"\n완료: {path}")
